"""Disease harmonizer QA graph — loads app_graph TSVs and builds Cytoscape payloads."""
from __future__ import annotations

import csv
import re
import io
import json
from difflib import SequenceMatcher
from collections import Counter, OrderedDict, defaultdict
from pathlib import Path
from typing import Any, TYPE_CHECKING

from fastapi import HTTPException

if TYPE_CHECKING:
    from src.qa_browser.target_id_graph import TargetGraphData


class DiseaseGraphData:
    """Indexes built from the four app_graph TSV files."""

    __slots__ = (
        "concepts_by_pxref",
        "concepts_by_ncats_id",
        "edges_by_pxref",
        "hierarchy_by_child",
        "hierarchy_by_parent",
        "clinical_descendants_by_pxref",
        "labels_by_xref_id",
        "decisions_by_pxref",
        "associations_by_ncats_id",
        "manifest",
        "source_catalog",
        "_dashboard_stats",
        "_xref_to_pxrefs",
        "_resolver_terms",
    )

    def __init__(self) -> None:
        self.concepts_by_pxref: dict[str, dict] = {}
        self.concepts_by_ncats_id: dict[str, dict] = {}
        self.edges_by_pxref: dict[str, list[dict]] = defaultdict(list)
        self.hierarchy_by_child: dict[str, list[dict]] = defaultdict(list)
        self.hierarchy_by_parent: dict[str, list[dict]] = defaultdict(list)
        self.clinical_descendants_by_pxref: dict[str, list[dict]] = defaultdict(list)
        self.labels_by_xref_id: dict[str, dict] = {}
        self.decisions_by_pxref: dict[str, list[dict]] = defaultdict(list)
        self.associations_by_ncats_id: dict[str, list[dict]] = defaultdict(list)
        self.manifest: dict = {}
        self.source_catalog: list[dict[str, str]] = []
        self._dashboard_stats: dict | None = None
        self._xref_to_pxrefs: dict[str, set[str]] | None = None
        self._resolver_terms: list[dict[str, Any]] | None = None


_singleton: DiseaseGraphData | None = None


def _read_tsv(path: Path) -> list[dict[str, str]]:
    with open(path, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh, delimiter="\t"))


def _source_catalog_candidates(data_dir: Path) -> list[Path]:
    candidates = [data_dir / "disease_source_catalog.tsv"]
    if data_dir.name == "app_graph":
        candidates.append(data_dir.parent / "metadata" / "disease_source_catalog.tsv")
    else:
        candidates.append(data_dir / "metadata" / "disease_source_catalog.tsv")
        candidates.append(data_dir.parent / "metadata" / "disease_source_catalog.tsv")
    seen: set[Path] = set()
    unique: list[Path] = []
    for candidate in candidates:
        if candidate not in seen:
            seen.add(candidate)
            unique.append(candidate)
    return unique


def _load_source_catalog(data_dir: Path) -> list[dict[str, str]]:
    for path in _source_catalog_candidates(data_dir):
        if path.exists():
            return _read_tsv(path)
    return []


_RESOLVED_DECISION_STATUSES = {"resolved", "acknowledged", "wontfix"}

REVIEW_DECISION_OPTIONS = [
    {"value": "Accept as-is", "label": "Accept as-is"},
    {"value": "Accept consensus", "label": "Accept consensus"},
    {"value": "Accept source", "label": "Accept source"},
    {"value": "Remove obsolete xref", "label": "Remove obsolete xref"},
    {"value": "Omit xref", "label": "Omit xref"},
    {"value": "Demote from exact", "label": "Demote from exact"},
    {"value": "Use replacement", "label": "Use replacement"},
    {"value": "Do not merge / split concern", "label": "Do not merge / split concern"},
    {"value": "needs_expert_review", "label": "needs_expert_review"},
    {"value": "Needs upstream ticket", "label": "Needs upstream ticket"},
]

REVIEW_INTAKE_COLUMNS = [
    "Registry ID",
    "Review decision",
    "Corrected xref / replacement",
    "Reviewer notes",
    "Primary Xref",
    "Disease Name",
    "NCATS Disease ID",
    "Review category",
    "Xref ID",
    "Xref Namespace",
    "Decision type",
    "Scenario ID",
    "Source value",
    "Consensus value",
    "Evidence note",
    "Reviewed by",
    "Reviewed at",
    "App review ID",
]

_OPEN_REVIEW_STATUSES = {"", "open"}


def normalize_review_text(value: Any) -> str:
    """Normalize old user-facing review vocabulary for app display.

    NOTE: Keep in sync with disease_app_graph_export._review_display_text
    """
    text = str(value or "")
    legacy_program = "R" + "DIP"
    replacements = {
        f"Needs {legacy_program}/domain review": "needs_expert_review",
        f"{legacy_program}/domain review": "needs_expert_review",
        f"{legacy_program}/GARD review": "needs_expert_review",
        f"domain/{legacy_program} review": "needs_expert_review",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def _join_unique(values: list[str] | tuple[str, ...]) -> str:
    seen: set[str] = set()
    unique: list[str] = []
    for value in values:
        text = str(value).strip()
        if text and text not in seen:
            seen.add(text)
            unique.append(text)
    return "|".join(unique)


def _nodenorm_status(concept: dict) -> str:
    """Support both old app_graph and current TargetGraph NodeNorm column names."""
    return (
        concept.get("nodenorm_validation_status", "")
        or concept.get("nodenorm_concordance", "")
    )


def _decision_action(decision: dict) -> str:
    return decision.get("auto_decision", "") or decision.get("resolution", "")


def _decision_display_type(decision: dict) -> str:
    """Prefer the curator/QC action over raw divergence labels for resolved rows."""
    raw_type = decision.get("decision_type", "")
    status = decision.get("status", "")
    action = _decision_action(decision)
    action_key = action.lower().replace(" ", "_")
    if status in _RESOLVED_DECISION_STATUSES:
        if action_key in {"omit_obsolete_xref", "remove_obsolete_xref"}:
            return "resolved_obsolete_xref"
        if action_key:
            return f"resolved_{action_key}"
    return raw_type


def _decision_graph_group_key(decision: dict) -> tuple[str, ...]:
    status = decision.get("status", "")
    if status in _RESOLVED_DECISION_STATUSES:
        return (
            "resolved",
            decision.get("xref_id", ""),
            _decision_display_type(decision),
            status,
            _decision_action(decision),
        )
    return ("raw", decision.get("decision_id", ""))


def _graph_id_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.:-]+", "_", str(value).strip() or "none")


def _merge_decision_group(pxref: str, decisions: list[dict]) -> dict:
    first = dict(decisions[0])
    ids = [d.get("decision_id", "") for d in decisions]
    display_type = _decision_display_type(first)
    status = first.get("status", "")
    action = _decision_action(first)
    if len(decisions) > 1:
        first["decision_id"] = "::".join([
            "merged_decision",
            _graph_id_part(pxref),
            _graph_id_part(first.get("xref_id", "")),
            _graph_id_part(display_type),
            _graph_id_part(status),
            _graph_id_part(action),
        ])
    first["source_decision_ids"] = _join_unique(ids)
    first["decision_type"] = _join_unique([d.get("decision_type", "") for d in decisions])
    first["decision_source"] = _join_unique([d.get("decision_source", "") for d in decisions])
    first["scenario_id"] = _join_unique([d.get("scenario_id", "") for d in decisions])
    first["auto_confidence"] = _join_unique([d.get("auto_confidence", "") for d in decisions])
    first["auto_rationale"] = _join_unique([d.get("auto_rationale", "") for d in decisions])
    first["resolution"] = _join_unique([d.get("resolution", "") for d in decisions])
    first["resolution_detail"] = _join_unique([d.get("resolution_detail", "") for d in decisions])
    first["reviewed_by"] = _join_unique([d.get("reviewed_by", "") for d in decisions])
    first["priority"] = _join_unique([d.get("priority", "") for d in decisions])
    return first


def _aggregate_graph_decisions(pxref: str, decisions: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, ...], list[dict]] = {}
    for decision in decisions:
        grouped.setdefault(_decision_graph_group_key(decision), []).append(decision)
    return [_merge_decision_group(pxref, group) for group in grouped.values()]


def _pick_edge_decision(decisions: list[dict]) -> dict:
    """Pick the decision most useful for the graph label."""
    if not decisions:
        return {}
    for decision in decisions:
        if decision.get("status", "") == "open":
            return decision
    for decision in decisions:
        action = _decision_action(decision)
        action_key = action.lower().replace(" ", "_")
        if action_key in {"omit_obsolete_xref", "remove_obsolete_xref"}:
            return decision
    return decisions[0]


def _load_app_graph_data(data_dir: Path) -> DiseaseGraphData:
    if not data_dir.is_dir():
        raise HTTPException(status_code=500, detail=f"Disease app_graph dir not found: {data_dir}")

    data = DiseaseGraphData()

    manifest_path = data_dir / "manifest.json"
    if manifest_path.exists():
        with open(manifest_path, encoding="utf-8") as fh:
            data.manifest = json.load(fh)
    data.source_catalog = _load_source_catalog(data_dir)

    for row in _read_tsv(data_dir / "disease_concepts.tsv"):
        pxref = row.get("primary_xref", "")
        ncats_id = row.get("ncats_disease_id", "")
        if pxref:
            data.concepts_by_pxref[pxref] = row
        if ncats_id:
            data.concepts_by_ncats_id[ncats_id] = row

    for row in _read_tsv(data_dir / "disease_xref_edges.tsv"):
        pxref = row.get("primary_xref", "")
        if pxref:
            data.edges_by_pxref[pxref].append(row)

    xref_labels_path = data_dir / "xref_labels.tsv"
    if xref_labels_path.exists():
        for row in _read_tsv(xref_labels_path):
            xref_id = row.get("xref_id", "")
            if xref_id:
                data.labels_by_xref_id[xref_id] = row

    review_decisions_path = data_dir / "review_decisions.tsv"
    if review_decisions_path.exists():
        for row in _read_tsv(review_decisions_path):
            pxref = row.get("primary_xref", "")
            if pxref:
                data.decisions_by_pxref[pxref].append(row)

    hierarchy_path = data_dir / "disease_hierarchy_edges.tsv"
    if hierarchy_path.exists():
        for row in _read_tsv(hierarchy_path):
            child = row.get("child_primary_xref", "")
            parent = row.get("parent_primary_xref", "")
            if child:
                data.hierarchy_by_child[child].append(row)
            if parent:
                data.hierarchy_by_parent[parent].append(row)

    clinical_descendants_path = data_dir / "disease_clinical_descendant_edges.tsv"
    if clinical_descendants_path.exists():
        for row in _read_tsv(clinical_descendants_path):
            pxref = row.get("primary_xref", "")
            if pxref:
                data.clinical_descendants_by_pxref[pxref].append(row)

    gene_assoc_path = data_dir / "disease_gene_associations.tsv"
    if gene_assoc_path.exists():
        for row in _read_tsv(gene_assoc_path):
            ncats_id = row.get("ncats_disease_id", "")
            if ncats_id:
                data.associations_by_ncats_id[ncats_id].append(row)

    # Validate loaded row counts against manifest
    if data.manifest:
        files_meta = data.manifest.get("files", {})
        checks = [
            ("disease_concepts.tsv", len(data.concepts_by_pxref),
             files_meta.get("disease_concepts.tsv", {}).get("rows")),
            ("disease_xref_edges.tsv", sum(len(v) for v in data.edges_by_pxref.values()),
             files_meta.get("disease_xref_edges.tsv", {}).get("rows")),
        ]
        for fname, actual, expected in checks:
            if expected is not None and actual != expected:
                print(f"WARNING: {fname} loaded {actual:,} rows, manifest says {expected:,}")

    return data


def _print_graph_load(label: str, data: DiseaseGraphData) -> None:
    n_concepts = len(data.concepts_by_pxref)
    n_edges = sum(len(v) for v in data.edges_by_pxref.values())
    n_hierarchy = sum(len(v) for v in data.hierarchy_by_child.values())
    n_clinical_descendants = sum(len(v) for v in data.clinical_descendants_by_pxref.values())
    n_labels = len(data.labels_by_xref_id)
    n_decisions = sum(len(v) for v in data.decisions_by_pxref.values())
    n_gene_assoc = sum(len(v) for v in data.associations_by_ncats_id.values())
    print(
        f"{label} loaded: {n_concepts:,} concepts, {n_edges:,} xref edges, "
        f"{n_hierarchy:,} hierarchy edges, "
        f"{n_clinical_descendants:,} clinical descendant edges, "
        f"{n_labels:,} labels, {n_decisions:,} decisions, "
        f"{n_gene_assoc:,} gene associations"
    )


def load_disease_graph_data(data_dir: str | Path, *, force_reload: bool = False) -> DiseaseGraphData:
    """Load TSVs + manifest, build indexes, cache as module singleton."""
    global _singleton
    if _singleton is not None and not force_reload:
        return _singleton

    data_dir = Path(data_dir)
    data = _load_app_graph_data(data_dir)
    _singleton = data
    _print_graph_load("Disease graph", data)
    return data


def reload_disease_graph_data(data_dir: str | Path) -> DiseaseGraphData:
    """Force-reload disease graph data, clearing the cached singleton."""
    return load_disease_graph_data(data_dir, force_reload=True)


def compute_dashboard_stats(data: DiseaseGraphData) -> dict[str, Any]:
    """Aggregate dashboard stats, computed once and cached on the singleton."""
    if data._dashboard_stats is not None:
        return data._dashboard_stats

    total_concepts = len(data.concepts_by_pxref)
    total_edges = sum(len(v) for v in data.edges_by_pxref.values())
    total_hierarchy_edges = sum(len(v) for v in data.hierarchy_by_child.values())
    total_clinical_descendant_edges = sum(
        len(v) for v in data.clinical_descendants_by_pxref.values()
    )
    total_labels = len(data.labels_by_xref_id)
    total_decisions = sum(len(v) for v in data.decisions_by_pxref.values())

    confidence_dist: Counter[str] = Counter()
    quality_dist: Counter[str] = Counter()
    edge_evidence_tier_dist: Counter[str] = Counter()
    edge_review_tier_dist: Counter[str] = Counter()
    edge_relation_type_dist: Counter[str] = Counter()
    source_coverage: Counter[str] = Counter()
    nodenorm_dist: Counter[str] = Counter()
    triage_dist: Counter[str] = Counter()
    disease_type_dist: Counter[str] = Counter()
    rare_count = 0
    flagged_count = 0
    qc_decision_count = 0
    auto_cleared_count = 0
    multi_source_count = 0
    source_only_count = 0
    total_xrefs = 0
    concepts_with_clinical_codes = 0
    concepts_with_clinical_descendants = 0

    # Per-source-namespace quality breakdown
    _sqb_tier: dict[str, Counter[str]] = {}
    _sqb_quality_sum: Counter[str] = Counter()
    _sqb_quality_n: Counter[str] = Counter()
    _sqb_needs_action: Counter[str] = Counter()
    _sqb_total: Counter[str] = Counter()

    clinical_descendant_source_dist: Counter[str] = Counter()
    clinical_descendant_distance_dist: Counter[str] = Counter()
    clinical_descendant_warning_dist: Counter[str] = Counter()
    clinical_descendant_review_dist: Counter[str] = Counter()

    for edges in data.edges_by_pxref.values():
        for edge in edges:
            relation_type = (edge.get("relation_type") or edge.get("match_type") or "unclassified").strip()
            evidence_tier = (edge.get("evidence_tier") or "unknown").strip()
            review_tier = (edge.get("review_tier") or "unknown").strip()
            edge_relation_type_dist[relation_type or "unknown"] += 1
            edge_evidence_tier_dist[evidence_tier or "unknown"] += 1
            edge_review_tier_dist[review_tier or "unknown"] += 1

    for pxref, concept in data.concepts_by_pxref.items():
        tier = concept.get("confidence_tier", "").strip()
        tier_key = tier if tier else "unknown"
        confidence_dist[tier_key] += 1

        quality = concept.get("overall_quality", "").strip()
        quality_dist[quality if quality else "unknown"] += 1

        nn_status = _nodenorm_status(concept).strip()
        nodenorm_dist[nn_status if nn_status else "unknown"] += 1

        disease_type = concept.get("disease_type", "").strip()
        disease_type_dist[disease_type if disease_type else "unknown"] += 1

        triage_bucket = concept.get("needs_review_triage_bucket", "").strip()
        if triage_bucket and triage_bucket != "not_needs_review":
            triage_dist[triage_bucket] += 1
        if concept.get("needs_review_auto_cleared", "").lower() == "true":
            auto_cleared_count += 1
        if _needs_review_decision(concept):
            qc_decision_count += 1

        if concept.get("is_rare", "").lower() == "true":
            rare_count += 1

        n_sources = _safe_int(concept.get("n_sources"))
        if n_sources > 1:
            multi_source_count += 1
        else:
            source_only_count += 1

        xc = int(concept.get("xref_count") or 0)
        total_xrefs += xc

        # Action-needed concepts only. Historical QC decisions remain visible
        # separately through needs_review_decision/evidence_note.
        decisions = data.decisions_by_pxref.get(pxref, [])
        is_flagged = _is_flagged(concept, decisions)
        if is_flagged:
            flagged_count += 1

        # Source namespace coverage from edges + per-source quality tracking
        quality_val = 0.0
        try:
            quality_val = float(quality) if quality else 0.0
        except (ValueError, TypeError):
            pass
        seen_ns: set[str] = set()
        has_clinical_code_xref = False
        for edge in data.edges_by_pxref.get(pxref, []):
            ns = _normalize_ns(edge.get("xref_namespace", "").strip())
            if ns:
                seen_ns.add(ns)
                if ns in {"ICD10", "ICD10CM", "ICD11", "ICD11f", "SNOMEDCT"}:
                    has_clinical_code_xref = True
        if has_clinical_code_xref:
            concepts_with_clinical_codes += 1
        for ns in seen_ns:
            source_coverage[ns] += 1
            _sqb_total[ns] += 1
            if ns not in _sqb_tier:
                _sqb_tier[ns] = Counter()
            _sqb_tier[ns][tier_key] += 1
            _sqb_quality_sum[ns] += quality_val
            _sqb_quality_n[ns] += 1
            if is_flagged:
                _sqb_needs_action[ns] += 1

        clinical_rows = data.clinical_descendants_by_pxref.get(pxref, [])
        if clinical_rows:
            concepts_with_clinical_descendants += 1
        for descendant in clinical_rows:
            clinical_descendant_source_dist[
                descendant.get("hierarchy_source", "").strip() or "unknown"
            ] += 1
            clinical_descendant_distance_dist[
                descendant.get("distance_from_anchor", "").strip()
                or descendant.get("distance_status", "").strip()
                or "unknown"
            ] += 1
            warning_key = (
                "may_not_be_rare"
                if descendant.get("may_not_be_rare_warning", "").lower() == "true"
                else "no_warning"
            )
            clinical_descendant_warning_dist[warning_key] += 1
            clinical_descendant_review_dist[
                descendant.get("review_status", "").strip() or "unreviewed"
            ] += 1

    avg_xrefs = round(total_xrefs / total_concepts, 1) if total_concepts else 0
    multi_source_percent = round((multi_source_count / total_concepts) * 100, 1) if total_concepts else 0

    # Gene association stats
    total_gene_associations = sum(len(v) for v in data.associations_by_ncats_id.values())
    unique_genes: set[str] = set()
    concepts_with_genes = 0
    gene_tier_dist: Counter[str] = Counter()
    gene_source_coverage: Counter[str] = Counter()
    for ncats_id, assocs in data.associations_by_ncats_id.items():
        if assocs:
            concepts_with_genes += 1
        for assoc in assocs:
            symbol = assoc.get("subject_symbol", "").strip()
            if symbol:
                unique_genes.add(symbol)
            tier = assoc.get("support_tier", "").strip()
            if tier:
                gene_tier_dist[tier] += 1
            for src in assoc.get("supporting_sources", "").split("|"):
                src = src.strip()
                if src:
                    gene_source_coverage[src] += 1

    # Build per-source quality breakdown
    source_quality_breakdown: dict[str, dict[str, Any]] = {}
    for ns in sorted(_sqb_total):
        avg_q = round(_sqb_quality_sum[ns] / _sqb_quality_n[ns], 2) if _sqb_quality_n[ns] else 0.0
        source_quality_breakdown[ns] = {
            "tier_counts": dict(_sqb_tier.get(ns, Counter()).most_common()),
            "avg_quality": avg_q,
            "needs_action": _sqb_needs_action.get(ns, 0),
            "total": _sqb_total[ns],
        }

    stats = {
        "total_concepts": total_concepts,
        "total_edges": total_edges,
        "total_hierarchy_edges": total_hierarchy_edges,
        "total_clinical_descendant_edges": total_clinical_descendant_edges,
        "total_labels": total_labels,
        "total_decisions": total_decisions,
        "confidence_distribution": dict(confidence_dist.most_common()),
        "quality_distribution": dict(quality_dist.most_common()),
        "edge_evidence_tier_distribution": dict(edge_evidence_tier_dist.most_common()),
        "edge_review_tier_distribution": dict(edge_review_tier_dist.most_common()),
        "edge_relation_type_distribution": dict(edge_relation_type_dist.most_common()),
        "source_coverage": dict(source_coverage.most_common()),
        "nodenorm_distribution": dict(nodenorm_dist.most_common()),
        "disease_type_distribution": dict(disease_type_dist.most_common()),
        "triage_distribution": dict(triage_dist.most_common()),
        "rare_count": rare_count,
        "flagged_count": flagged_count,
        "needs_action_count": flagged_count,
        "needs_review_count": confidence_dist.get("needs_review", 0),
        "qc_decision_count": qc_decision_count,
        "auto_cleared_count": auto_cleared_count,
        "multi_source_count": multi_source_count,
        "multi_source_percent": multi_source_percent,
        "source_only_count": source_only_count,
        "avg_xrefs_per_concept": avg_xrefs,
        "xref_density_label": "Average direct xrefs per concept; coverage context only, not confidence.",
        "concepts_with_clinical_codes": concepts_with_clinical_codes,
        "concepts_with_clinical_descendants": concepts_with_clinical_descendants,
        "clinical_descendant_summary": {
            "total": total_clinical_descendant_edges,
            "concepts": concepts_with_clinical_descendants,
            "by_hierarchy_source": dict(clinical_descendant_source_dist.most_common()),
            "by_distance": dict(clinical_descendant_distance_dist.most_common()),
            "by_warning": dict(clinical_descendant_warning_dist.most_common()),
            "by_review_status": dict(clinical_descendant_review_dist.most_common()),
        },
        "source_quality_breakdown": source_quality_breakdown,
        "total_gene_associations": total_gene_associations,
        "unique_genes": len(unique_genes),
        "concepts_with_genes": concepts_with_genes,
        "gene_support_tier_distribution": dict(gene_tier_dist.most_common()),
        "gene_source_coverage": dict(gene_source_coverage.most_common()),
    }
    data._dashboard_stats = stats
    return stats


def parse_disease_ids(raw: str) -> list[str]:
    """Parse user input into a list of primary_xref or ncats_disease_id strings."""
    ids: list[str] = []
    for token in (raw or "").replace("|", " ").replace(",", " ").split():
        token = token.strip()
        if token and token not in ids:
            ids.append(token)
    return ids


def _ensure_xref_index(data: DiseaseGraphData) -> dict[str, set[str]]:
    """Build (or return cached) reverse index from xref_id -> set of primary_xrefs."""
    if data._xref_to_pxrefs is None:
        rev: dict[str, set[str]] = defaultdict(set)
        for px, edges in data.edges_by_pxref.items():
            for e in edges:
                xid = e.get("xref_id", "")
                if xid:
                    rev[xid].add(px)
        data._xref_to_pxrefs = dict(rev)
    return data._xref_to_pxrefs


def _resolve_to_pxref(data: DiseaseGraphData, query_id: str) -> str | None:
    """Resolve a query ID to a primary_xref. Accepts primary_xref, ncats_disease_id, or xref_id."""
    if query_id in data.concepts_by_pxref:
        return query_id
    concept = data.concepts_by_ncats_id.get(query_id)
    if concept:
        return concept.get("primary_xref")
    # Fallback: check xref reverse index (unambiguous single match only)
    xref_to_pxrefs = _ensure_xref_index(data)
    pxrefs = xref_to_pxrefs.get(query_id, set())
    if len(pxrefs) == 1:
        return next(iter(pxrefs))
    return None


def _concept_type_class(concept: dict) -> str:
    disease_type = concept.get("disease_type", "")
    if disease_type == "biolink:PhenotypicFeature":
        return "phenotype-concept"
    if disease_type == "biolink:DiseaseOrPhenotypicFeature":
        return "mixed-disease-phenotype-concept"
    return "disease-concept"


def _concept_classes(concept: dict, classes: str = "concept") -> str:
    parts = [part for part in classes.split() if part]
    parts.append(_concept_type_class(concept))
    return " ".join(dict.fromkeys(parts))


def _concept_graph_node(data: DiseaseGraphData, pxref: str, classes: str = "concept") -> dict | None:
    """Build one Cytoscape concept node."""
    concept = data.concepts_by_pxref.get(pxref)
    if concept is None:
        return None
    concept_node_id = f"concept::{pxref}"
    standard_name = concept.get("standard_name", "")
    concept_label = f"{pxref}\n{standard_name}" if standard_name else pxref
    return {
        "data": {
            "id": concept_node_id,
            "label": concept_label,
            "kind": "concept",
            "standard_name": concept.get("standard_name", ""),
            "ncats_disease_id": concept.get("ncats_disease_id", ""),
            "confidence_tier": concept.get("confidence_tier", ""),
            "confidence_tier_original": concept.get("confidence_tier_original", ""),
            "needs_review_auto_cleared": concept.get("needs_review_auto_cleared", ""),
            "needs_review_triage_bucket": concept.get("needs_review_triage_bucket", ""),
            "needs_review_decision": _needs_review_decision(concept),
            "overall_quality": concept.get("overall_quality", ""),
            "n_sources": concept.get("n_sources", ""),
            "primary_sources": concept.get("primary_sources", ""),
            "disease_type": concept.get("disease_type", ""),
            "is_rare": concept.get("is_rare", ""),
            "xref_count": concept.get("xref_count", ""),
            "cardinality_issue_count": concept.get("cardinality_issue_count", ""),
            "obsolete_xref_count": concept.get("obsolete_xref_count", ""),
            "nodenorm_canonical_curie": concept.get("nodenorm_canonical_curie", ""),
            "hierarchy_parent_count": str(len(data.hierarchy_by_child.get(pxref, []))),
            "hierarchy_child_count": str(len(data.hierarchy_by_parent.get(pxref, []))),
        },
        "classes": _concept_classes(concept, classes),
    }


def _extract_primary_predicate(predicates_str: str) -> str:
    """Pick the primary biolink predicate from a pipe-delimited string."""
    if not predicates_str or predicates_str == "nan":
        return ""
    parts = [p.strip() for p in predicates_str.split("|") if p.strip()]
    for part in parts:
        if part.startswith("biolink:"):
            return part
    return parts[0] if parts else ""


def _short_predicate(predicate: str) -> str:
    """Strip ``biolink:`` prefix for display labels."""
    return predicate.replace("biolink:", "") if predicate else ""


def _resolve_target_id_safe(target_data: "TargetGraphData | None", symbol: str) -> str | None:
    """Resolve a gene symbol to a target_id, returning None on failure."""
    if target_data is None:
        return None
    try:
        from src.qa_browser.target_id_graph import _resolve_target_id
        return _resolve_target_id(target_data, symbol)
    except Exception:
        return None


def _target_relation(
    source_id: str,
    target_id: str,
    predicate: str,
    edge_row: dict | None = None,
) -> dict:
    """Build a Cytoscape edge element for a target harmonizer relation."""
    edge_id = f"target-rel::{source_id}::{target_id}::{predicate}"
    label = _short_predicate(predicate)
    classes = "target-relation-edge"
    if predicate == "biolink:has_gene_product":
        classes += " gene-product-edge"
    elif predicate == "biolink:transcribed_to":
        classes += " transcription-edge"
    elif predicate == "biolink:translates_to":
        classes += " translation-edge"
    data: dict[str, Any] = {
        "id": edge_id,
        "source": source_id,
        "target": target_id,
        "label": label,
        "kind": "target_relation_edge",
        "predicate": predicate,
    }
    if edge_row:
        data["relation_kind"] = edge_row.get("relation_kind", "")
        data["evidence_identifier"] = edge_row.get("evidence_identifier", "")
        data["evidence_namespace"] = edge_row.get("evidence_namespace", "")
    return {"data": data, "classes": classes}


def build_disease_graph_payload(
    data: DiseaseGraphData,
    query_ids: list[str],
    target_data: "TargetGraphData | None" = None,
) -> dict[str, Any]:
    """Build Cytoscape elements for the given disease IDs."""
    if not query_ids:
        raise HTTPException(status_code=400, detail="Provide at least one disease ID.")

    concept_nodes: dict[str, dict] = {}
    xref_nodes: dict[str, dict] = {}
    decision_nodes: dict[str, dict] = {}
    gene_nodes: dict[str, dict] = {}
    gene_transcripts: dict[str, list[dict]] = {}
    gene_proteins: dict[str, list[dict]] = {}
    gene_product_counts: dict[str, dict[str, int]] = {}
    edges: dict[str, dict] = {}
    missing: list[str] = []
    namespaces: set[str] = set()
    seen_target_products: set[str] = set()

    for qid in query_ids:
        pxref = _resolve_to_pxref(data, qid)
        if pxref is None:
            missing.append(qid)
            continue

        concept = data.concepts_by_pxref[pxref]
        concept_node_id = f"concept::{pxref}"
        standard_name = concept.get("standard_name", "")
        concept_label = f"{pxref}\n{standard_name}" if standard_name else pxref
        concept_nodes[pxref] = {
            "data": {
                "id": concept_node_id,
                "label": concept_label,
                "kind": "concept",
                "standard_name": concept.get("standard_name", ""),
                "ncats_disease_id": concept.get("ncats_disease_id", ""),
                "confidence_tier": concept.get("confidence_tier", ""),
                "confidence_tier_original": concept.get("confidence_tier_original", ""),
                "needs_review_auto_cleared": concept.get("needs_review_auto_cleared", ""),
                "needs_review_triage_bucket": concept.get("needs_review_triage_bucket", ""),
                "needs_review_decision": _needs_review_decision(concept),
                "overall_quality": concept.get("overall_quality", ""),
                "n_sources": concept.get("n_sources", ""),
                "primary_sources": concept.get("primary_sources", ""),
                "disease_type": concept.get("disease_type", ""),
                "is_rare": concept.get("is_rare", ""),
                "xref_count": concept.get("xref_count", ""),
                "cardinality_issue_count": concept.get("cardinality_issue_count", ""),
                "obsolete_xref_count": concept.get("obsolete_xref_count", ""),
                "nodenorm_canonical_curie": concept.get("nodenorm_canonical_curie", ""),
                "hierarchy_parent_count": str(len(data.hierarchy_by_child.get(pxref, []))),
                "hierarchy_child_count": str(len(data.hierarchy_by_parent.get(pxref, []))),
            },
            "classes": _concept_classes(concept, "concept"),
        }

        # --- build decision lookup keyed by xref_id for this concept ---
        _dec_by_xref: dict[str, list[dict]] = defaultdict(list)
        for dec in data.decisions_by_pxref.get(pxref, []):
            dxid = dec.get("xref_id", "")
            if dxid:
                _dec_by_xref[dxid].append(dec)

        # --- xref edges ---
        for edge_row in data.edges_by_pxref.get(pxref, []):
            xref_id = edge_row.get("xref_id", "")
            if not xref_id:
                continue
            ns = _normalize_ns(edge_row.get("xref_namespace", "")).lower()
            namespaces.add(ns)
            xref_node_id = f"xref::{xref_id}"
            label_row = data.labels_by_xref_id.get(xref_id, {})
            raw_label = edge_row.get("xref_label", "") or label_row.get("preferred_label", "")
            pref_label = "" if raw_label in ("nan", "NaN") else raw_label
            xref_display = f"{xref_id}\n{pref_label}" if pref_label else xref_id
            xref_nodes.setdefault(xref_id, {
                "data": {
                    "id": xref_node_id,
                    "label": xref_display,
                    "kind": "xref",
                    "xref_namespace": _normalize_ns(edge_row.get("xref_namespace", "")),
                    "xref_label": pref_label,
                    "is_obsolete": label_row.get("is_obsolete", ""),
                    "obsolete_detail": label_row.get("obsolete_detail", ""),
                    "label_source": label_row.get("label_source", ""),
                    "lookup_status": label_row.get("lookup_status", ""),
                },
                "classes": f"xref {ns}",
            })

            match_type = edge_row.get("match_type", "unclassified")
            relation_type = edge_row.get("relation_type", "") or match_type
            label_sim = edge_row.get("label_similarity", "")
            xref_conf = edge_row.get("xref_confidence", "")
            equiv_conf = edge_row.get("equivalence_confidence", "")
            rel_conf = edge_row.get("relation_confidence", "")
            evidence_tier = edge_row.get("evidence_tier", "")
            label_score = equiv_conf or rel_conf or xref_conf
            if label_score and label_score != "0.0":
                edge_label = f"{relation_type} [eq={label_score}]"
            elif label_sim and label_sim != "0.0":
                edge_label = f"{relation_type} ({label_sim})"
            else:
                edge_label = relation_type
            if evidence_tier:
                edge_label += f" {evidence_tier}"

            # --- enrich edge with conflict/resolution data ---
            edge_decs = _dec_by_xref.get(xref_id, [])
            has_conflict = ""
            conflict_status = ""
            conflict_type = ""
            raw_conflict_type = ""
            conflict_resolution = ""
            conflict_detail = ""
            edge_classes = f"match-{match_type}"

            if edge_decs:
                has_conflict = "true"
                best = _pick_edge_decision(edge_decs)
                matching_decs = [
                    d for d in edge_decs
                    if _decision_graph_group_key(d) == _decision_graph_group_key(best)
                ]
                dec_status = best.get("status", "")
                dec_type = _join_unique([d.get("decision_type", "") for d in matching_decs])
                dec_display_type = _decision_display_type(best)
                dec_resolution = _decision_action(best)
                dec_detail = _join_unique([
                    d.get("auto_rationale", "") or d.get("resolution_detail", "")
                    for d in matching_decs
                ])
                conflict_status = (
                    dec_status if dec_status in _RESOLVED_DECISION_STATUSES else "open"
                )
                conflict_type = dec_display_type
                raw_conflict_type = dec_type
                conflict_resolution = dec_resolution
                conflict_detail = dec_detail
                edge_classes += " has-conflict"

                # Append conflict info to edge label
                if conflict_status == "open":
                    edge_label += " \u26a0 " + (dec_display_type or "conflict")
                else:
                    short_res = dec_resolution.replace("_", " ") if dec_resolution else "resolved"
                    # Truncate long resolutions for the label
                    if len(short_res) > 20:
                        short_res = short_res[:18] + "\u2026"
                    edge_label += f" \u26a0 resolved: {short_res}"
                    edge_classes += " conflict-resolved"

            edge_id = f"edge::{pxref}::{xref_id}"
            edges[edge_id] = {
                "data": {
                    "id": edge_id,
                    "source": concept_node_id,
                    "target": xref_node_id,
                    "label": edge_label,
                    "kind": "xref_edge",
                    "match_type": match_type,
                    "relation_type": relation_type,
                    "match_type_source": edge_row.get("match_type_source", ""),
                    "source_asserted": edge_row.get("source_asserted", ""),
                    "agreement_level": edge_row.get("agreement_level", ""),
                    "confidence_score": edge_row.get("confidence_score", ""),
                    "label_similarity": label_sim,
                    "mapping_confidence": edge_row.get("mapping_confidence", ""),
                    "label_confidence": edge_row.get("label_confidence", ""),
                    "source_support": edge_row.get("source_support", ""),
                    "validation_support": edge_row.get("validation_support", ""),
                    "corroboration_score": edge_row.get("corroboration_score", ""),
                    "asserting_sources": edge_row.get("asserting_sources", ""),
                    "validation_sources": edge_row.get("validation_sources", ""),
                    "evidence_streams": edge_row.get("evidence_streams", ""),
                    "structural_confidence": edge_row.get("structural_confidence", ""),
                    "xref_confidence": xref_conf,
                    "relation_confidence": rel_conf,
                    "equivalence_confidence": equiv_conf,
                    "evidence_tier": evidence_tier,
                    "review_tier": edge_row.get("review_tier", ""),
                    "score_rationale": edge_row.get("score_rationale", ""),
                    "xref_is_obsolete": edge_row.get("xref_is_obsolete", ""),
                    "override_status": edge_row.get("override_status", ""),
                    "override_reason": edge_row.get("override_reason", ""),
                    "has_conflict": has_conflict,
                    "conflict_status": conflict_status,
                    "conflict_type": conflict_type,
                    "raw_conflict_type": raw_conflict_type,
                    "conflict_resolution": conflict_resolution,
                    "conflict_detail": conflict_detail,
                },
                "classes": edge_classes,
            }

        # --- decisions ---
        for dec in _aggregate_graph_decisions(pxref, data.decisions_by_pxref.get(pxref, [])):
            dec_id = dec.get("decision_id", "")
            if not dec_id:
                continue
            dec_node_id = f"decision::{dec_id}"
            dec_type = dec.get("decision_type", "")
            dec_status = dec.get("status", "")
            dec_resolution = _decision_action(dec)
            dec_display_type = _decision_display_type(dec)
            dec_label_type = dec_display_type.replace("_", " ")
            # Build a descriptive label: "conflict [open]" or "conflict [resolved: downgrade_match_type]"
            if dec_resolution:
                dec_label = f"{dec_label_type}\n[{dec_status}: {dec_resolution}]"
            else:
                dec_label = f"{dec_label_type}\n[{dec_status}]"
            dec_xref_id = dec.get("xref_id", "")
            decision_nodes[dec_id] = {
                "data": {
                    "id": dec_node_id,
                    "label": dec_label,
                    "kind": "decision",
                    "decision_id": dec_id,
                    "source_decision_ids": dec.get("source_decision_ids", dec_id),
                    "decision_source": dec.get("decision_source", ""),
                    "decision_type": dec_display_type,
                    "raw_decision_type": dec_type,
                    "scenario_id": dec.get("scenario_id", ""),
                    "status": dec_status,
                    "auto_decision": dec_resolution,
                    "auto_confidence": dec.get("auto_confidence", ""),
                    "auto_rationale": dec.get("auto_rationale", ""),
                    "resolution": dec.get("resolution", ""),
                    "resolution_detail": dec.get("resolution_detail", ""),
                    "reviewed_by": dec.get("reviewed_by", ""),
                    "xref_id": dec_xref_id,
                    "priority": dec.get("priority", ""),
                },
                "classes": f"decision decision-{dec_status}",
            }
            # Connect decision to its xref node (if the xref is in the graph)
            if dec_xref_id:
                xref_target_id = f"xref::{dec_xref_id}"
                dec_edge_id = f"dec-edge::{dec_id}::{dec_xref_id}"
                edges[dec_edge_id] = {
                    "data": {
                        "id": dec_edge_id,
                        "source": dec_node_id,
                        "target": xref_target_id,
                        "label": dec_resolution or dec_status,
                        "kind": "decision_edge",
                    },
                    "classes": "decision-edge",
                }

        # --- gene associations ---
        ncats_id = concept.get("ncats_disease_id", "")
        for assoc in data.associations_by_ncats_id.get(ncats_id, []):
            symbol = assoc.get("subject_symbol", "").strip()
            if not symbol:
                continue
            gene_node_id = f"gene::{symbol}"
            tier = assoc.get("support_tier", "unknown")
            n_sources = assoc.get("n_supporting_sources", "")
            gene_public_id = assoc.get("gene_public_id", "")
            primary_predicate = _extract_primary_predicate(assoc.get("predicates", ""))

            # Enriched gene label: "SYMBOL\nHGNC:xxxxx"
            gene_label = f"{symbol}\n{gene_public_id}" if gene_public_id else symbol

            # Resolve symbol to target harmonizer
            target_tid = _resolve_target_id_safe(target_data, symbol)
            gene_classes = "gene"
            gene_data: dict[str, Any] = {
                "id": gene_node_id,
                "label": gene_label,
                "kind": "gene",
                "ncats_gene_id": assoc.get("ncats_gene_id", ""),
                "gene_public_id": gene_public_id,
                "subject_symbol": symbol,
                "source_subject_ids": assoc.get("source_subject_ids", ""),
            }

            if target_tid and target_data is not None:
                gene_classes = "gene target-linked"
                tnode = target_data.nodes_by_id.get(target_tid, {})
                gene_data["target_id"] = target_tid
                gene_data["target_name"] = tnode.get("target_name", "")
                gene_data["target_primary_id"] = tnode.get("primary_id", "")
                gene_data["target_ids"] = tnode.get("ids", "")
                gene_data["target_mapping_ratio"] = tnode.get("mapping_ratio", "")

            gene_nodes.setdefault(symbol, {
                "data": gene_data,
                "classes": gene_classes,
            })

            # Enriched association edge label
            short_pred = _short_predicate(primary_predicate)
            assoc_label = f"{short_pred}\n[{tier}, {n_sources} src]" if short_pred else f"{tier} [{n_sources}]"

            assoc_edge_id = f"assoc::{pxref}::{symbol}"
            edges[assoc_edge_id] = {
                "data": {
                    "id": assoc_edge_id,
                    "source": concept_node_id,
                    "target": gene_node_id,
                    "label": assoc_label,
                    "kind": "association_edge",
                    "subject_symbol": symbol,
                    "ncats_gene_id": assoc.get("ncats_gene_id", ""),
                    "gene_public_id": gene_public_id,
                    "primary_predicate": primary_predicate,
                    "support_tier": tier,
                    "support_score_raw": assoc.get("support_score_raw", ""),
                    "n_supporting_sources": n_sources,
                    "supporting_sources": assoc.get("supporting_sources", ""),
                    "curated_support_count": assoc.get("curated_support_count", ""),
                    "predicates": assoc.get("predicates", ""),
                    "evidence_types": assoc.get("evidence_types", ""),
                    "evidence_labels": assoc.get("evidence_labels", ""),
                    "publications": assoc.get("publications", ""),
                    "source_subject_ids": assoc.get("source_subject_ids", ""),
                    "source_disease_ids": assoc.get("source_disease_ids", ""),
                    "primary_knowledge_sources": assoc.get("primary_knowledge_sources", ""),
                    "aggregator_knowledge_sources": assoc.get("aggregator_knowledge_sources", ""),
                    "provided_bys": assoc.get("provided_bys", ""),
                },
                "classes": f"association-edge support-{_graph_id_part(tier)}",
            }

            # --- collect target products for on-demand expansion ---
            if target_tid and target_data is not None and symbol not in gene_transcripts and symbol not in gene_proteins:
                # Step 1: Collect transcripts and proteins from gene edges
                candidate_transcripts: list[tuple[dict, str]] = []  # (product_el, tid)
                candidate_proteins: list[tuple[dict, str, dict]] = []  # (product_el, tid, gene_edge)
                seen_tids: set[str] = set()
                for rel in target_data.relation_edges_by_target.get(target_tid, []):
                    rel_predicate = rel.get("predicate", "")
                    if rel_predicate not in ("biolink:has_gene_product", "biolink:transcribed_to"):
                        continue
                    rel_source = rel.get("source_id", "")
                    rel_target = rel.get("target_id", "")
                    product_tid = rel_target if rel_source == target_tid else rel_source
                    if product_tid in seen_target_products or product_tid in seen_tids:
                        continue
                    seen_tids.add(product_tid)
                    product_node = target_data.nodes_by_id.get(product_tid)
                    if not product_node:
                        continue
                    product_type = product_node.get("target_type", "")
                    product_node_id = f"target-product::{product_tid}"
                    product_primary_id = product_node.get("primary_id", "")
                    product_label = product_primary_id or product_node.get("symbol", product_tid)
                    if product_type == "protein":
                        p_classes = "target-product protein"
                    elif product_type == "transcript":
                        p_classes = "target-product transcript"
                    else:
                        p_classes = "target-product"
                    product_el = {
                        "data": {
                            "id": product_node_id,
                            "label": product_label,
                            "kind": "target_product",
                            "target_type": product_type,
                            "target_id": product_tid,
                            "target_name": product_node.get("target_name", ""),
                            "target_primary_id": product_primary_id,
                            "target_ids": product_node.get("ids", ""),
                        },
                        "classes": p_classes,
                    }
                    if product_type == "protein":
                        candidate_proteins.append((product_el, product_tid, rel))
                    elif product_type == "transcript":
                        candidate_transcripts.append((product_el, product_tid))

                # Step 2: Build translates_to index (protein_tid → (transcript_tid, edge_row))
                transcript_tid_set = {tid for _, tid in candidate_transcripts}
                translates_to_map: dict[str, tuple[str, dict]] = {}
                for _, p_tid, _ in candidate_proteins:
                    for tt_edge in target_data.relation_edges_by_target.get(p_tid, []):
                        if tt_edge.get("predicate") != "biolink:translates_to":
                            continue
                        src = tt_edge.get("source_id", "")
                        tgt = tt_edge.get("target_id", "")
                        tt_tid = src if tgt == p_tid else tgt
                        if tt_tid in transcript_tid_set:
                            translates_to_map[p_tid] = (tt_tid, tt_edge)
                            break

                # Step 3: Cap selection — up to 5 each, overflow fills other slots
                max_products = 10
                max_per_type = 5
                transcripts_take = candidate_transcripts[:max_per_type]
                proteins_take = candidate_proteins[:max_per_type]
                remaining = max_products - len(proteins_take) - len(transcripts_take)
                if remaining > 0 and len(candidate_proteins) > max_per_type:
                    proteins_take.extend(candidate_proteins[max_per_type:max_per_type + remaining])
                    remaining = max_products - len(proteins_take) - len(transcripts_take)
                if remaining > 0 and len(candidate_transcripts) > max_per_type:
                    transcripts_take.extend(candidate_transcripts[max_per_type:max_per_type + remaining])

                # Step 4: Build separate transcript and protein element lists
                selected_transcript_tids = {tid for _, tid in transcripts_take}
                t_elements: list[dict] = []
                p_elements: list[dict] = []

                for product_el, t_tid in transcripts_take:
                    seen_target_products.add(t_tid)
                    t_elements.append(product_el)
                    t_elements.append(_target_relation(
                        gene_node_id, product_el["data"]["id"],
                        "biolink:transcribed_to",
                    ))

                for product_el, p_tid, gene_rel in proteins_take:
                    seen_target_products.add(p_tid)
                    product_el["data"]["subject_symbol"] = symbol
                    p_elements.append(product_el)
                    # Use translates_to when the translating transcript is
                    # in the transcript set; otherwise fall back to
                    # has_gene_product from the gene node.
                    if p_tid in translates_to_map:
                        tt_tid, tt_edge = translates_to_map[p_tid]
                        if tt_tid in selected_transcript_tids:
                            transcript_node_id = f"target-product::{tt_tid}"
                            p_elements.append(_target_relation(
                                transcript_node_id, product_el["data"]["id"],
                                "biolink:translates_to", tt_edge,
                            ))
                            # Fallback edge for when transcripts are hidden
                            fallback = _target_relation(
                                gene_node_id, product_el["data"]["id"],
                                "biolink:has_gene_product", gene_rel,
                            )
                            fallback["data"]["_fallback"] = True
                            p_elements.append(fallback)
                            continue
                    p_elements.append(_target_relation(
                        gene_node_id, product_el["data"]["id"],
                        "biolink:has_gene_product", gene_rel,
                    ))

                total_proteins = len(candidate_proteins)
                total_transcripts = len(candidate_transcripts)
                if t_elements:
                    gene_transcripts[symbol] = t_elements
                if p_elements:
                    gene_proteins[symbol] = p_elements
                gene_product_counts[symbol] = {
                    "transcripts": total_transcripts,
                    "proteins": total_proteins,
                }
                gene_nodes[symbol]["data"]["target_protein_count"] = str(total_proteins)
                gene_nodes[symbol]["data"]["target_transcript_count"] = str(total_transcripts)
                gene_nodes[symbol]["data"]["target_product_count"] = str(
                    total_proteins + total_transcripts
                )

    elements = [
        *concept_nodes.values(),
        *xref_nodes.values(),
        *decision_nodes.values(),
        *gene_nodes.values(),
        *edges.values(),
    ]

    result: dict[str, Any] = {
        "queryIds": query_ids,
        "missingIds": missing,
        "namespaces": sorted(namespaces),
        "stats": {
            "conceptCount": len(concept_nodes),
            "xrefCount": len(xref_nodes),
            "namespaceCount": len(namespaces),
            "edgeCount": len(edges),
            "decisionCount": len(decision_nodes),
            "geneCount": len(gene_nodes),
        },
        "manifest": data.manifest,
        "elements": elements,
    }
    if gene_transcripts:
        result["geneTranscripts"] = gene_transcripts
    if gene_proteins:
        result["geneProteins"] = gene_proteins
    if gene_product_counts:
        result["geneProductCounts"] = gene_product_counts
    return result


def build_hierarchy_graph_payload(data: DiseaseGraphData, query_id: str) -> dict[str, Any]:
    """Build immediate MONDO parent/child context for one concept."""
    pxref = _resolve_to_pxref(data, query_id)
    if pxref is None:
        raise HTTPException(status_code=404, detail=f"Concept not found: {query_id}")

    concept_nodes: dict[str, dict] = {}
    edges: dict[str, dict] = {}

    def add_node(node_pxref: str, extra_classes: str) -> None:
        node = _concept_graph_node(data, node_pxref, f"concept {extra_classes}".strip())
        if node is None:
            return
        if node_pxref in concept_nodes:
            existing = set(concept_nodes[node_pxref].get("classes", "").split())
            existing.update(node.get("classes", "").split())
            concept_nodes[node_pxref]["classes"] = " ".join(sorted(existing))
        else:
            concept_nodes[node_pxref] = node

    add_node(pxref, "hierarchy-focus")

    for row in data.hierarchy_by_child.get(pxref, []):
        parent = row.get("parent_primary_xref", "")
        child = row.get("child_primary_xref", pxref)
        if not parent or not child:
            continue
        add_node(parent, "hierarchy-context hierarchy-parent")
        add_node(child, "hierarchy-focus")
        edge_id = f"hier-edge::{child}::{parent}"
        edges[edge_id] = {
            "data": {
                "id": edge_id,
                "source": f"concept::{child}",
                "target": f"concept::{parent}",
                "label": "subclass_of",
                "kind": "hierarchy_edge",
                "relationship_type": row.get("relationship_type", ""),
                "predicate": row.get("predicate", ""),
                "hierarchy_source": row.get("source", ""),
                "parent_primary_xref": parent,
                "parent_label": row.get("parent_label", ""),
                "child_primary_xref": child,
                "child_label": row.get("child_label", ""),
            },
            "classes": "hierarchy-edge hierarchy-parent-edge",
        }

    for row in data.hierarchy_by_parent.get(pxref, []):
        parent = row.get("parent_primary_xref", pxref)
        child = row.get("child_primary_xref", "")
        if not parent or not child:
            continue
        add_node(parent, "hierarchy-focus")
        add_node(child, "hierarchy-context hierarchy-child")
        edge_id = f"hier-edge::{child}::{parent}"
        edges[edge_id] = {
            "data": {
                "id": edge_id,
                "source": f"concept::{child}",
                "target": f"concept::{parent}",
                "label": "subclass_of",
                "kind": "hierarchy_edge",
                "relationship_type": row.get("relationship_type", ""),
                "predicate": row.get("predicate", ""),
                "hierarchy_source": row.get("source", ""),
                "parent_primary_xref": parent,
                "parent_label": row.get("parent_label", ""),
                "child_primary_xref": child,
                "child_label": row.get("child_label", ""),
            },
            "classes": "hierarchy-edge hierarchy-child-edge",
        }

    elements = [*concept_nodes.values(), *edges.values()]
    return {
        "queryIds": [query_id],
        "missingIds": [],
        "stats": {
            "conceptCount": len(concept_nodes),
            "hierarchyEdgeCount": len(edges),
            "parentCount": len(data.hierarchy_by_child.get(pxref, [])),
            "childCount": len(data.hierarchy_by_parent.get(pxref, [])),
        },
        "manifest": data.manifest,
        "elements": elements,
    }


def build_clinical_descendant_graph_payload(
    data: DiseaseGraphData,
    query_id: str,
    limit: int = 80,
) -> dict[str, Any]:
    """Build clinical-code descendant candidate context for one concept."""
    pxref = _resolve_to_pxref(data, query_id)
    if pxref is None:
        raise HTTPException(status_code=404, detail=f"Concept not found: {query_id}")

    concept = data.concepts_by_pxref.get(pxref)
    concept_node = _concept_graph_node(data, pxref, "concept clinical-descendant-focus")
    elements: list[dict[str, Any]] = []
    if concept_node:
        elements.append(concept_node)

    rows = data.clinical_descendants_by_pxref.get(pxref, [])[:limit]
    for row in rows:
        descendant_xref = row.get("descendant_xref_id", "")
        if not descendant_xref:
            continue
        ns = _normalize_ns(row.get("descendant_xref_namespace", ""))
        ns_class = ns.lower()
        xref_node_id = f"clinical-descendant::{descendant_xref}"
        desc_label = row.get("descendant_xref_label", "")
        display_label = f"{descendant_xref}\n{desc_label}" if desc_label else descendant_xref
        elements.append({
            "data": {
                "id": xref_node_id,
                "label": display_label,
                "kind": "clinical_descendant_xref",
                "xref_namespace": ns,
                "xref_label": desc_label,
                "relationship_to_disease": row.get("relationship_to_disease", ""),
                "descendant_source_status": row.get("descendant_source_status", ""),
                "may_not_be_rare_warning": row.get("may_not_be_rare_warning", ""),
                "warning_reason": row.get("warning_reason", ""),
            },
            "classes": f"xref clinical-descendant-xref {ns_class}",
        })

        edge_id = "::".join([
            "clinical-descendant-edge",
            pxref,
            row.get("anchor_xref_id", ""),
            descendant_xref,
        ])
        distance = row.get("distance_from_anchor", "")
        label = "narrow candidate"
        if distance:
            label = f"{label} d={distance}"
        elements.append({
            "data": {
                "id": edge_id,
                "source": f"concept::{pxref}",
                "target": xref_node_id,
                "label": label,
                "kind": "clinical_descendant_edge",
                "match_type": "narrow",
                "predicate": row.get("predicate", ""),
                "relationship_to_anchor": row.get("relationship_to_anchor", ""),
                "relationship_to_disease": row.get("relationship_to_disease", ""),
                "hierarchy_source": row.get("hierarchy_source", ""),
                "distance_from_anchor": distance,
                "distance_status": row.get("distance_status", ""),
                "path": row.get("path", ""),
                "anchor_xref_id": row.get("anchor_xref_id", ""),
                "anchor_xref_namespace": row.get("anchor_xref_namespace", ""),
                "anchor_xref_label": row.get("anchor_xref_label", ""),
                "anchor_match_type": row.get("anchor_match_type", ""),
                "anchor_asserting_sources": row.get("anchor_asserting_sources", ""),
                "descendant_xref_id": descendant_xref,
                "descendant_xref_namespace": ns,
                "descendant_xref_label": desc_label,
                "descendant_source_status": row.get("descendant_source_status", ""),
                "existing_odin_match_types": row.get("existing_odin_match_types", ""),
                "existing_odin_primary_xrefs": row.get("existing_odin_primary_xrefs", ""),
                "in_mondo_xrefs": row.get("in_mondo_xrefs", ""),
                "in_yadaw_s2": row.get("in_yadaw_s2", ""),
                "in_yadaw_s3": row.get("in_yadaw_s3", ""),
                "may_not_be_rare_warning": row.get("may_not_be_rare_warning", ""),
                "warning_reason": row.get("warning_reason", ""),
                "evidence_sources": row.get("evidence_sources", ""),
                "review_status": row.get("review_status", ""),
            },
            "classes": "clinical-descendant-edge match-narrow",
        })

    return {
        "queryIds": [query_id],
        "missingIds": [],
        "stats": {
            "conceptCount": 1 if concept else 0,
            "clinicalDescendantCount": len(rows),
        },
        "manifest": data.manifest,
        "elements": elements,
    }


def load_flagged_concepts(data: DiseaseGraphData, limit: int = 500) -> list[dict[str, Any]]:
    """Return concepts that still need review/action for the landing page table."""
    flagged: list[dict[str, Any]] = []
    for pxref, concept in data.concepts_by_pxref.items():
        cardinality = int(concept.get("cardinality_issue_count") or 0)
        obsolete = int(concept.get("obsolete_xref_count") or 0)
        has_obsolete = concept.get("has_obsolete", "").lower() == "true"

        if not _is_flagged(concept, data.decisions_by_pxref.get(pxref, [])):
            continue

        flags: list[str] = []
        if cardinality > 0:
            flags.append(f"cardinality({cardinality})")
        if obsolete > 0 or has_obsolete:
            flags.append(f"obsolete({obsolete})")
        _append_needs_review_flags(concept, flags)

        flagged.append({
            "primary_xref": pxref,
            "standard_name": concept.get("standard_name", ""),
            "confidence_tier": concept.get("confidence_tier", ""),
            "confidence_tier_original": concept.get("confidence_tier_original", ""),
            "needs_review_auto_cleared": concept.get("needs_review_auto_cleared", ""),
            "needs_review_triage_bucket": concept.get("needs_review_triage_bucket", ""),
            "needs_review_triage_action": concept.get("needs_review_triage_action", ""),
            "needs_review_problem_namespaces": concept.get("needs_review_problem_namespaces", ""),
            "needs_review_decision": _needs_review_decision(concept),
            "evidence_note": concept.get("evidence_note", ""),
            "overall_quality": concept.get("overall_quality", ""),
            "n_sources": concept.get("n_sources", ""),
            "xref_count": concept.get("xref_count", ""),
            "flags": ", ".join(flags),
            "href": f"/disease-id-qa?ids={pxref}",
        })
        if len(flagged) >= limit:
            break

    return flagged


def export_flagged_tsv(data: DiseaseGraphData, limit: int = 5000) -> str:
    """Return action-needed concepts as a TSV string for download."""
    rows = load_flagged_concepts(data, limit=limit)
    if not rows:
        return ""
    columns = [
        "primary_xref", "standard_name", "confidence_tier",
        "confidence_tier_original", "needs_review_auto_cleared",
        "needs_review_triage_bucket", "needs_review_triage_action",
        "needs_review_problem_namespaces", "needs_review_decision",
        "evidence_note", "overall_quality", "n_sources", "xref_count", "flags",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, delimiter="\t", extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def _review_category(decision: dict, concept: dict) -> str:
    """Bucket a review decision into a compact UI category."""
    text = " ".join([
        decision.get("decision_type", ""),
        decision.get("scenario_id", ""),
        decision.get("auto_decision", ""),
        decision.get("auto_rationale", ""),
        decision.get("resolution", ""),
        concept.get("needs_review_triage_bucket", ""),
        concept.get("needs_review_triage_action", ""),
        concept.get("evidence_note", ""),
    ]).lower()
    tier = concept.get("confidence_tier", "").lower()
    namespace = decision.get("xref_namespace", "").upper()
    if "obsolete" in text or tier == "review_obsolete_xref":
        return "obsolete_xref"
    if "validator" in text or tier == "review_validator_disagreement":
        return "validator_disagreement"
    if "broad" in text or "narrow" in text or "split" in text:
        return "broad_narrow_or_split"
    if "phenotype" in text or concept.get("disease_type", "") == "biolink:PhenotypicFeature":
        return "phenotype_vs_disease"
    if "cardinality" in text or _safe_int(concept.get("cardinality_issue_count")) > 0:
        return "cardinality_issue"
    if namespace:
        return f"{namespace.lower()}_conflict"
    if "conflict" in text or "dissent" in text:
        return "source_conflict"
    return "needs_review"


def _review_category_label(category: str) -> str:
    labels = {
        "obsolete_xref": "Obsolete xref",
        "validator_disagreement": "Validator disagreement",
        "broad_narrow_or_split": "Broad/narrow or split concern",
        "phenotype_vs_disease": "Phenotype vs disease",
        "cardinality_issue": "Cardinality issue",
        "source_conflict": "Source conflict",
        "needs_review": "Needs review",
    }
    if category.endswith("_conflict"):
        return f"{category[:-9].upper()} conflict"
    return labels.get(category, category.replace("_", " "))


def _review_status_filter(status: str) -> set[str] | None:
    status = (status or "open").strip().lower()
    if status in {"all", "any"}:
        return None
    if status in {"open", "unresolved", "needs_action"}:
        return _OPEN_REVIEW_STATUSES
    if status in {"stale", "inactive"}:
        return {"stale"}
    if status in {"resolved", "closed"}:
        return _RESOLVED_DECISION_STATUSES
    return {status}


def _review_item_status(item: dict[str, Any]) -> str:
    return (item.get("status", "") or "open").strip().lower()


def _review_xref_tokens(value: str) -> list[str]:
    return [token.strip() for token in (value or "").split("|") if token.strip()]


def _build_review_edge_indexes(
    data: DiseaseGraphData,
) -> tuple[set[tuple[str, str, str]], dict[str, set[str]]]:
    edge_keys: set[tuple[str, str, str]] = set()
    xref_to_pxrefs: dict[str, set[str]] = defaultdict(set)
    for pxref, edges in data.edges_by_pxref.items():
        for edge in edges:
            xref_id = edge.get("xref_id", "")
            ns = _normalize_ns(edge.get("xref_namespace", ""))
            if not xref_id:
                continue
            edge_keys.add((pxref, xref_id, ns))
            xref_to_pxrefs[xref_id].add(pxref)
    return edge_keys, xref_to_pxrefs


_STALE_REASON_LABELS = {
    "edge_still_present": "Edge still present; old QC conflict no longer reproduced",
    "edge_split_or_changed": "Old combined xref assertion is now split or changed",
    "xref_attached_elsewhere": "Xref now attached to a different primary concept",
    "xref_edge_absent": "Xref edge no longer present in current graph",
    "primary_concept_absent": "Primary concept no longer present in current graph",
}


def _stale_review_reason(
    data: DiseaseGraphData,
    item: dict[str, Any],
    edge_keys: set[tuple[str, str, str]],
    xref_to_pxrefs: dict[str, set[str]],
) -> tuple[str, str]:
    pxref = item.get("primary_xref", "")
    ns = _normalize_ns(item.get("xref_namespace", ""))
    xref_tokens = _review_xref_tokens(item.get("xref_id", ""))
    if pxref not in data.concepts_by_pxref:
        key = "primary_concept_absent"
    elif any((pxref, xref_id, ns) in edge_keys for xref_id in xref_tokens):
        key = "edge_still_present"
    elif len(xref_tokens) > 1 and any(
        any((pxref, xref_id, edge_ns) in edge_keys for edge_ns in ("", ns))
        or any(candidate_px == pxref for candidate_px in xref_to_pxrefs.get(xref_id, set()))
        for xref_id in xref_tokens
    ):
        key = "edge_split_or_changed"
    elif any(xref_id in xref_to_pxrefs for xref_id in xref_tokens):
        key = "xref_attached_elsewhere"
    else:
        key = "xref_edge_absent"
    return key, _STALE_REASON_LABELS[key]


def _review_item_from_decision(
    pxref: str,
    concept: dict,
    decision: dict,
) -> dict[str, Any]:
    category = _review_category(decision, concept)
    rationale = (
        decision.get("auto_rationale", "")
        or decision.get("resolution", "")
        or concept.get("evidence_note", "")
    )
    registry_id = decision.get("decision_id", "")
    return {
        "review_id": registry_id or f"concept::{pxref}",
        "registry_id": registry_id,
        "saveable": bool(registry_id and decision.get("decision_source") == "divergence_registry"),
        "decision_source": decision.get("decision_source", ""),
        "primary_xref": pxref,
        "ncats_disease_id": concept.get("ncats_disease_id", ""),
        "standard_name": concept.get("standard_name", ""),
        "category": category,
        "category_label": _review_category_label(category),
        "xref_id": decision.get("xref_id", ""),
        "xref_namespace": _normalize_ns(decision.get("xref_namespace", "")),
        "decision_type": decision.get("decision_type", ""),
        "scenario_id": decision.get("scenario_id", ""),
        "status": decision.get("status", ""),
        "auto_decision": decision.get("auto_decision", ""),
        "auto_confidence": decision.get("auto_confidence", ""),
        "source_value": decision.get("source_value", ""),
        "consensus_value": decision.get("consensus_value", ""),
        "n_sources_agree": decision.get("n_sources_agree", ""),
        "n_sources_total": decision.get("n_sources_total", ""),
        "evidence_note": normalize_review_text(rationale),
        "reviewed_by": decision.get("reviewed_by", ""),
        "date_found": decision.get("date_found", ""),
        "date_resolved": decision.get("date_resolved", ""),
        "graph_href": f"/disease-id-qa?ids={pxref}&tab=graph",
        "provenance_href": f"/disease-id-qa/api/provenance/{pxref}",
    }


def _review_item_from_concept(pxref: str, concept: dict) -> dict[str, Any]:
    decision = {
        "decision_id": "",
        "decision_source": "concept_triage",
        "primary_xref": pxref,
        "ncats_disease_id": concept.get("ncats_disease_id", ""),
        "standard_name": concept.get("standard_name", ""),
        "xref_id": "",
        "xref_namespace": concept.get("needs_review_problem_namespaces", ""),
        "decision_type": concept.get("needs_review_triage_bucket", ""),
        "scenario_id": "",
        "status": "open",
        "auto_decision": concept.get("needs_review_triage_action", ""),
        "auto_confidence": "",
        "auto_rationale": concept.get("evidence_note", ""),
        "source_value": "",
        "consensus_value": concept.get("authority_consensus", ""),
        "n_sources_agree": "",
        "n_sources_total": concept.get("n_sources", ""),
    }
    return _review_item_from_decision(pxref, concept, decision)


def _iter_review_items(
    data: DiseaseGraphData,
    *,
    include_resolved: bool = False,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for pxref, concept in data.concepts_by_pxref.items():
        decisions = data.decisions_by_pxref.get(pxref, [])
        active_decisions = [
            decision for decision in decisions
            if decision.get("status", "").strip().lower() not in _RESOLVED_DECISION_STATUSES
        ]
        row_decisions = decisions if include_resolved else active_decisions
        for decision in row_decisions:
            items.append(_review_item_from_decision(pxref, concept, decision))
        if not active_decisions and _is_flagged(concept, decisions):
            items.append(_review_item_from_concept(pxref, concept))
    return items


def build_review_queue(
    data: DiseaseGraphData,
    *,
    category: str = "",
    source: str = "",
    status: str = "open",
    q: str = "",
    page: int = 1,
    per_page: int = 50,
) -> dict[str, Any]:
    """Return clustered review items with pagination for the app Review tab."""
    q_lower = q.strip().lower()
    wanted_status = _review_status_filter(status)
    source_upper = source.strip().upper()

    all_items = _iter_review_items(data, include_resolved=True)
    status_counts: Counter[str] = Counter(_review_item_status(item) for item in all_items)
    edge_keys, xref_to_pxrefs = _build_review_edge_indexes(data)
    stale_reason_source_counts: dict[str, Counter[str]] = defaultdict(Counter)
    stale_reason_key_counts: Counter[str] = Counter()
    obsolete_omitted_count = 0
    for item in all_items:
        if _review_item_status(item) == "stale":
            reason_key, reason_label = _stale_review_reason(
                data,
                item,
                edge_keys,
                xref_to_pxrefs,
            )
            item["stale_reason"] = reason_label
            item["stale_reason_key"] = reason_key
            stale_reason_key_counts[reason_key] += 1
            ns = item.get("xref_namespace", "")
            if ns:
                stale_reason_source_counts[reason_key][ns] += 1
        action = (
            item.get("auto_decision", "")
            or item.get("resolution", "")
            or ""
        ).lower().replace(" ", "_")
        if (
            _review_item_status(item) in _RESOLVED_DECISION_STATUSES
            and action in {"omit_obsolete_xref", "remove_obsolete_xref"}
        ):
            obsolete_omitted_count += 1

    review_summary = {
        "still_needs_decision": sum(status_counts.get(s, 0) for s in _OPEN_REVIEW_STATUSES),
        "stale_or_inactive": status_counts.get("stale", 0),
        "reviewed_or_resolved": sum(status_counts.get(s, 0) for s in _RESOLVED_DECISION_STATUSES),
        "obsolete_omitted": obsolete_omitted_count,
        "all_review_records": len(all_items),
        "stale_reasons": [
            {
                "reason": key,
                "label": _STALE_REASON_LABELS.get(key, key),
                "count": count,
            }
            for key, count in stale_reason_key_counts.most_common()
        ],
        "stale_reason_counts": dict(stale_reason_key_counts.most_common()),
        "stale_sources_by_reason": {
            key: dict(counts.most_common())
            for key, counts in stale_reason_source_counts.items()
        },
    }

    status_scoped_items = [
        item for item in all_items
        if wanted_status is None or _review_item_status(item) in wanted_status
    ]
    category_counts: Counter[str] = Counter(
        item["category"] for item in status_scoped_items
    )
    source_counts: Counter[str] = Counter(
        item["xref_namespace"] for item in status_scoped_items if item["xref_namespace"]
    )

    filtered: list[dict[str, Any]] = []
    for item in status_scoped_items:
        if category and item.get("category") != category:
            continue
        if source_upper and item.get("xref_namespace", "").upper() != source_upper:
            continue
        if q_lower:
            haystack = " ".join([
                item.get("primary_xref", ""),
                item.get("ncats_disease_id", ""),
                item.get("standard_name", ""),
                item.get("xref_id", ""),
                item.get("xref_namespace", ""),
                item.get("decision_type", ""),
                item.get("evidence_note", ""),
            ]).lower()
            if q_lower not in haystack:
                continue
        filtered.append(item)

    filtered.sort(
        key=lambda row: (
            row.get("category_label", ""),
            row.get("xref_namespace", ""),
            row.get("standard_name", "").lower(),
            row.get("primary_xref", ""),
        )
    )
    total = len(filtered)
    per_page = max(1, min(per_page, 200))
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    rows = filtered[start:start + per_page]

    return {
        "rows": rows,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "category": category,
        "source": source,
        "status": status,
        "query": q,
        "category_counts": [
            {"category": key, "label": _review_category_label(key), "count": count}
            for key, count in category_counts.most_common()
        ],
        "source_counts": dict(source_counts.most_common()),
        "status_counts": dict(status_counts.most_common()),
        "review_summary": review_summary,
        "decision_options": REVIEW_DECISION_OPTIONS,
    }


def export_review_intake_template(
    data: DiseaseGraphData,
    *,
    category: str = "",
    source: str = "",
    status: str = "open",
    q: str = "",
    limit: int = 5000,
) -> str:
    """Export a TargetGraph-compatible review intake TSV template."""
    queue = build_review_queue(
        data,
        category=category,
        source=source,
        status=status,
        q=q,
        page=1,
        per_page=min(max(1, limit), 5000),
    )
    rows = []
    for item in queue["rows"]:
        rows.append({
            "Registry ID": item.get("registry_id", ""),
            "Review decision": "",
            "Corrected xref / replacement": "",
            "Reviewer notes": "",
            "Primary Xref": item.get("primary_xref", ""),
            "Disease Name": item.get("standard_name", ""),
            "NCATS Disease ID": item.get("ncats_disease_id", ""),
            "Review category": item.get("category", ""),
            "Xref ID": item.get("xref_id", ""),
            "Xref Namespace": item.get("xref_namespace", ""),
            "Decision type": item.get("decision_type", ""),
            "Scenario ID": item.get("scenario_id", ""),
            "Source value": item.get("source_value", ""),
            "Consensus value": item.get("consensus_value", ""),
            "Evidence note": item.get("evidence_note", ""),
            "Reviewed by": "",
            "Reviewed at": "",
            "App review ID": item.get("review_id", ""),
        })
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=REVIEW_INTAKE_COLUMNS, delimiter="\t")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def _is_flagged(concept: dict, decisions: list[dict]) -> bool:
    """Return True if a concept still needs review/action.

    Do not treat decision provenance, original review status, or auto-cleared
    triage as flagged. In the UI, "flagged" should mean unresolved.
    """
    tier = concept.get("confidence_tier", "").strip().lower()
    if tier in {"needs_review", "review_validator_disagreement", "review_obsolete_xref"}:
        return True

    triage_bucket = concept.get("needs_review_triage_bucket", "").strip()
    triage_action = concept.get("needs_review_triage_action", "").strip()
    auto_cleared = concept.get("needs_review_auto_cleared", "").lower() == "true"
    return (
        bool(triage_bucket and triage_bucket != "not_needs_review")
        and triage_action == "manual_review"
        and not auto_cleared
    )


def _needs_review_decision(concept: dict) -> str:
    """Return the concept-level needs-review triage decision for UI/export."""
    bucket = concept.get("needs_review_triage_bucket", "").strip()
    if not bucket or bucket == "not_needs_review":
        return ""
    action = concept.get("needs_review_triage_action", "").strip()
    auto_cleared = concept.get("needs_review_auto_cleared", "").lower() == "true"
    prefix = "auto-cleared" if auto_cleared else "review-retained"
    if action:
        return f"{prefix}: {bucket} ({action})"
    return f"{prefix}: {bucket}"


def _append_needs_review_flags(concept: dict, flags: list[str]) -> None:
    """Append action-needed concept-level review flag text in-place."""
    original_tier = concept.get("confidence_tier_original", "").strip()
    current_tier = concept.get("confidence_tier", "").strip()
    bucket = concept.get("needs_review_triage_bucket", "").strip()
    auto_cleared = concept.get("needs_review_auto_cleared", "").lower() == "true"
    if current_tier == "needs_review":
        flags.append("needs_review")
    elif current_tier in {"review_validator_disagreement", "review_obsolete_xref"}:
        flags.append(current_tier)
    if bucket and bucket != "not_needs_review":
        if not auto_cleared:
            flags.append(f"review_retained({bucket})")
        elif original_tier == "needs_review" and current_tier == "needs_review":
            flags.append(f"auto_clear_inconsistent({bucket})")


def _concept_has_source(pxref: str, source: str, data: DiseaseGraphData) -> bool:
    """Return True if the concept has an xref edge with the given namespace."""
    source_upper = source.upper()
    for edge in data.edges_by_pxref.get(pxref, []):
        if edge.get("xref_namespace", "").upper() == source_upper:
            return True
    return False


def _format_hierarchy_refs(rows: list[dict], role: str, limit: int = 12) -> str:
    """Format a capped parent/child list for table rows and downloads."""
    id_key = f"{role}_primary_xref"
    label_key = f"{role}_label"
    values: list[str] = []
    for row in rows[:limit]:
        curie = row.get(id_key, "")
        label = row.get(label_key, "")
        if curie and label:
            values.append(f"{curie}: {label}")
        elif curie:
            values.append(curie)
    if len(rows) > limit:
        values.append(f"+{len(rows) - limit} more")
    return " | ".join(values)


_DISEASE_SOURCE_NAMESPACE_COLUMNS = [
    "DOID", "EFO", "GARD", "HP", "ICD10", "ICD11", "ICD11f",
    "ICD9", "ICDO", "KEGG", "MEDDRA", "MEDGEN", "MESH",
    "MONDO", "NCIT", "OMIM", "OMIMPS", "Orphanet", "SNOMEDCT", "UMLS",
]
_DISEASE_SOURCE_NAMESPACE_BY_UPPER = {
    ns.upper(): ns for ns in _DISEASE_SOURCE_NAMESPACE_COLUMNS
}
_DISEASE_SOURCE_NAMESPACE_ALIASES = {
    "ICD10CM": "ICD10",
    "ORPHA": "Orphanet",
    "ORDO": "Orphanet",
}


def _curie_namespace(curie: str) -> str:
    return curie.split(":", 1)[0].strip() if ":" in curie else ""


def _canonical_source_column(ns: str) -> str:
    normalized = _normalize_ns((ns or "").strip())
    key = normalized.upper()
    if key in _DISEASE_SOURCE_NAMESPACE_ALIASES:
        return _DISEASE_SOURCE_NAMESPACE_ALIASES[key]
    return _DISEASE_SOURCE_NAMESPACE_BY_UPPER.get(key, normalized)


def _append_unique(values: list[str], value: str) -> None:
    value = (value or "").strip()
    if value and value not in values:
        values.append(value)


def _concept_to_row(
    pxref: str,
    concept: dict,
    data: DiseaseGraphData,
    include_sources: set[str] | None = None,
    exclude_obsolete: bool = False,
) -> dict[str, Any]:
    """Convert a concept dict to a search result row.

    Parameters
    ----------
    include_sources : set[str] | None
        When set, only include xref edge labels whose namespace (upper-cased)
        is in this set.  ``None`` means include all sources.
    exclude_obsolete : bool
        When True, skip xref edges where ``xref_is_obsolete`` is true.
    """
    decisions = data.decisions_by_pxref.get(pxref, [])
    cardinality = int(concept.get("cardinality_issue_count") or 0)
    obsolete = int(concept.get("obsolete_xref_count") or 0)
    has_obsolete = concept.get("has_obsolete", "").lower() == "true"
    action_needed = _is_flagged(concept, decisions)

    flags: list[str] = []
    if action_needed and cardinality > 0:
        flags.append(f"cardinality({cardinality})")
    if action_needed and (obsolete > 0 or has_obsolete):
        flags.append(f"obsolete({obsolete})")
    _append_needs_review_flags(concept, flags)

    parents = data.hierarchy_by_child.get(pxref, [])
    children = data.hierarchy_by_parent.get(pxref, [])
    source_xrefs: dict[str, list[str]] = {
        ns: [] for ns in _DISEASE_SOURCE_NAMESPACE_COLUMNS
    }

    def add_source_xref(xref_id: str, namespace: str = "") -> None:
        column = _canonical_source_column(namespace or _curie_namespace(xref_id))
        if column not in source_xrefs:
            return
        if include_sources is not None and column.upper() not in include_sources:
            return
        _append_unique(source_xrefs[column], xref_id)

    add_source_xref(pxref)

    # Build compact source labels: "MONDO: Diabetes | OMIM: Type 2 diabetes"
    source_labels: list[str] = []
    for edge in data.edges_by_pxref.get(pxref, []):
        xref_id = edge.get("xref_id", "")
        ns = _normalize_ns(edge.get("xref_namespace", ""))
        if include_sources is not None and ns.upper() not in include_sources:
            continue
        if exclude_obsolete and edge.get("xref_is_obsolete", "").lower() in ("true", "1"):
            continue
        add_source_xref(xref_id, ns)
        label = edge.get("xref_label", "")
        if not label:
            label_row = data.labels_by_xref_id.get(xref_id, {})
            label = label_row.get("preferred_label", "")
        if ns and label:
            source_labels.append(f"{ns}: {label}")

    row = {
        "primary_xref": pxref,
        "ncats_disease_id": concept.get("ncats_disease_id", ""),
        "standard_name": concept.get("standard_name", ""),
        "confidence_tier": concept.get("confidence_tier", ""),
        "confidence_tier_original": concept.get("confidence_tier_original", ""),
        "confidence_tier_triaged": concept.get("confidence_tier_triaged", ""),
        "needs_review_auto_cleared": concept.get("needs_review_auto_cleared", ""),
        "needs_review_triage_bucket": concept.get("needs_review_triage_bucket", ""),
        "needs_review_triage_action": concept.get("needs_review_triage_action", ""),
        "needs_review_problem_namespaces": concept.get("needs_review_problem_namespaces", ""),
        "needs_review_cardinality_namespaces": concept.get("needs_review_cardinality_namespaces", ""),
        "needs_review_obsolete_namespaces": concept.get("needs_review_obsolete_namespaces", ""),
        "needs_review_validator_namespaces": (
            concept.get("needs_review_validator_namespaces", "")
            or concept.get("needs_review_discordant_namespaces", "")
        ),
        "needs_review_discordant_namespaces": concept.get("needs_review_discordant_namespaces", ""),
        "needs_review_decision": _needs_review_decision(concept),
        "authority_consensus": concept.get("authority_consensus", ""),
        "evidence_note": concept.get("evidence_note", ""),
        "overall_quality": concept.get("overall_quality", ""),
        "n_sources": concept.get("n_sources", ""),
        "xref_count": concept.get("xref_count", ""),
        "flags": ", ".join(flags),
        "source_labels": " | ".join(source_labels),
        "definition": concept.get("definition", ""),
        "synonyms": concept.get("synonyms", ""),
        "is_rare": concept.get("is_rare", ""),
        "disease_type": concept.get("disease_type", ""),
        "nodenorm_canonical_curie": concept.get("nodenorm_canonical_curie", ""),
        "nodenorm_canonical_label": concept.get("nodenorm_canonical_label", ""),
        "nodenorm_validation_status": _nodenorm_status(concept),
        "hierarchy_parent_count": len(parents),
        "hierarchy_child_count": len(children),
        "hierarchy_parents": _format_hierarchy_refs(parents, "parent"),
        "hierarchy_children": _format_hierarchy_refs(children, "child"),
        "href": f"/disease-id-qa?ids={pxref}",
    }
    row.update({
        ns: " | ".join(source_xrefs[ns])
        for ns in _DISEASE_SOURCE_NAMESPACE_COLUMNS
    })
    return row


def _split_pipe(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def _normalize_resolver_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text or text == "nan":
        return ""
    text = text.replace("'", "")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalize_resolver_id(value: Any) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"C\d{7}", text) or re.fullmatch(r"CN\d+", text):
        return f"UMLS:{text}"
    return text


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        text = str(value or "").strip()
        return float(text) if text else default
    except (TypeError, ValueError):
        return default


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        text = str(value or "").strip()
        return int(float(text)) if text else default
    except (TypeError, ValueError):
        return default


def _lexical_score(query_norm: str, term_norm: str) -> float:
    if not query_norm or not term_norm:
        return 0.0
    if query_norm == term_norm:
        return 1.0

    query_tokens = query_norm.split()
    term_tokens = term_norm.split()
    if not query_tokens or not term_tokens:
        return 0.0

    token_overlap = len(set(query_tokens) & set(term_tokens))
    containment = token_overlap / max(1, min(len(set(query_tokens)), len(set(term_tokens))))
    jaccard = token_overlap / max(1, len(set(query_tokens) | set(term_tokens)))
    sequence = SequenceMatcher(None, query_norm, term_norm).ratio()

    substring_score = 0.0
    if query_norm in term_norm or term_norm in query_norm:
        shorter = min(len(query_tokens), len(term_tokens))
        longer = max(len(query_tokens), len(term_tokens))
        substring_score = 0.82 + 0.12 * (shorter / max(1, longer))

    return round(max(sequence, containment * 0.96, jaccard * 0.9, substring_score), 4)


def _tier_score(tier: str) -> float:
    return {
        "multi_source_supported": 0.92,
        "single_authoritative_source": 0.86,
        "needs_review": 0.58,
        "review_validator_disagreement": 0.42,
        "review_obsolete_xref": 0.25,
    }.get((tier or "").strip(), 0.5)


def _field_prior(field: str) -> float:
    return {
        "primary_xref": 1.0,
        "ncats_disease_id": 1.0,
        "xref_id": 0.98,
        "standard_name": 0.96,
        "synonym": 0.93,
        "xref_label": 0.9,
        "xref_synonym": 0.86,
    }.get(field, 0.82)


def _relationship_to_query(term: dict[str, Any], lexical: float) -> str:
    if term.get("obsolete"):
        return "obsolete_identifier"
    match_type = term.get("match_type", "")
    if match_type in {"broad", "narrow", "related"}:
        return f"{match_type}_xref_match"
    if lexical >= 0.98:
        return "equivalent_candidate"
    return "lexical_candidate"


def _resolver_match_status(score: float, lexical: float) -> str:
    if lexical >= 0.98 and score >= 0.9:
        return "exact"
    if score >= 0.82:
        return "probable_exact"
    if score >= 0.68:
        return "possible"
    return "weak"


def _resolver_index(data: DiseaseGraphData) -> list[dict[str, Any]]:
    if data._resolver_terms is not None:
        return data._resolver_terms

    terms: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()

    def add_term(
        pxref: str,
        raw_term: str,
        field: str,
        *,
        source: str = "",
        xref_id: str = "",
        match_type: str = "",
        xref_confidence: str = "",
        equivalence_confidence: str = "",
        obsolete: bool = False,
    ) -> None:
        raw_term = str(raw_term or "").strip()
        if not raw_term or raw_term.lower() == "nan":
            return
        norm = _normalize_resolver_text(raw_term)
        key = (pxref, norm, field, xref_id)
        if not norm or key in seen:
            return
        seen.add(key)
        terms.append({
            "pxref": pxref,
            "term": raw_term,
            "norm": norm,
            "field": field,
            "source": source,
            "xref_id": xref_id,
            "match_type": match_type,
            "xref_confidence": xref_confidence,
            "equivalence_confidence": equivalence_confidence or xref_confidence,
            "obsolete": obsolete,
        })

    for pxref, concept in data.concepts_by_pxref.items():
        add_term(pxref, pxref, "primary_xref", source=concept.get("primary_xref_source", ""))
        add_term(pxref, concept.get("ncats_disease_id", ""), "ncats_disease_id")
        add_term(pxref, concept.get("standard_name", ""), "standard_name")
        for synonym in _split_pipe(concept.get("synonyms", "")):
            add_term(pxref, synonym, "synonym")

        for edge in data.edges_by_pxref.get(pxref, []):
            xref_id = edge.get("xref_id", "")
            ns = _normalize_ns(edge.get("xref_namespace", ""))
            is_obsolete = edge.get("xref_is_obsolete", "").lower() in {"true", "1"}
            is_obsolete = is_obsolete or edge.get("override_status", "") == "blocked"
            add_term(
                pxref,
                xref_id,
                "xref_id",
                source=ns,
                xref_id=xref_id,
                match_type=edge.get("match_type", ""),
                xref_confidence=edge.get("xref_confidence", ""),
                equivalence_confidence=edge.get("equivalence_confidence", ""),
                obsolete=is_obsolete,
            )
            add_term(
                pxref,
                edge.get("xref_label", ""),
                "xref_label",
                source=ns,
                xref_id=xref_id,
                match_type=edge.get("match_type", ""),
                xref_confidence=edge.get("xref_confidence", ""),
                equivalence_confidence=edge.get("equivalence_confidence", ""),
                obsolete=is_obsolete,
            )
            label_row = data.labels_by_xref_id.get(xref_id, {})
            add_term(
                pxref,
                label_row.get("preferred_label", ""),
                "xref_label",
                source=ns,
                xref_id=xref_id,
                match_type=edge.get("match_type", ""),
                xref_confidence=edge.get("xref_confidence", ""),
                equivalence_confidence=edge.get("equivalence_confidence", ""),
                obsolete=is_obsolete,
            )
            for synonym in _split_pipe(label_row.get("synonyms", "")):
                add_term(
                    pxref,
                    synonym,
                    "xref_synonym",
                    source=ns,
                    xref_id=xref_id,
                    match_type=edge.get("match_type", ""),
                    xref_confidence=edge.get("xref_confidence", ""),
                    equivalence_confidence=edge.get("equivalence_confidence", ""),
                    obsolete=is_obsolete,
                )

    data._resolver_terms = terms
    return terms


def resolve_name_candidates(
    data: DiseaseGraphData,
    query: str,
    limit: int = 10,
    include_obsolete: bool = False,
) -> dict[str, Any]:
    """Resolve a free-text disease name or CURIE against harmonized concepts."""
    query = str(query or "").strip()
    query_id = _normalize_resolver_id(query)
    query_norm = _normalize_resolver_text(query_id)
    if not query_norm:
        return {"query": query, "normalized_query": "", "candidates": [], "best": None}
    id_like_query = bool(
        re.fullmatch(r"[A-Za-z][A-Za-z0-9_.-]*:\S+", query_id)
        or re.fullmatch(r"UMLS:C\d{7}", query_id)
        or re.fullmatch(r"UMLS:CN\d+", query_id)
    )

    candidate_terms: dict[str, tuple[dict[str, Any], float]] = {}

    direct_pxref = _resolve_to_pxref(data, query_id)
    if direct_pxref:
        candidate_terms[direct_pxref] = ({
            "pxref": direct_pxref,
            "term": query_id,
            "norm": query_norm,
            "field": "primary_xref",
            "source": "primary",
            "xref_id": "",
            "match_type": "exact",
            "xref_confidence": "1.0",
            "equivalence_confidence": "1.0",
            "obsolete": False,
        }, 1.0)

    for term in _resolver_index(data):
        if term.get("obsolete") and not include_obsolete:
            continue
        if id_like_query:
            if term.get("field") not in {"primary_xref", "ncats_disease_id", "xref_id"}:
                continue
            if _normalize_resolver_id(term.get("term", "")).lower() != query_id.lower():
                continue
            lexical = 1.0
        else:
            lexical = 1.0 if query_id == term.get("term") else _lexical_score(query_norm, term.get("norm", ""))
        if lexical < 0.58:
            continue
        pxref = term["pxref"]
        current = candidate_terms.get(pxref)
        if current is None or lexical > current[1]:
            candidate_terms[pxref] = (term, lexical)

    candidates: list[dict[str, Any]] = []
    for pxref, (term, lexical) in candidate_terms.items():
        concept = data.concepts_by_pxref.get(pxref, {})
        source_count = int(_safe_float(concept.get("n_sources", 0)))
        source_score = min(1.0, 0.55 + source_count * 0.09)
        xref_score = _safe_float(
            term.get("equivalence_confidence", "") or term.get("xref_confidence", ""),
            0.0,
        )
        harmonizer_score = max(_tier_score(concept.get("confidence_tier", "")), xref_score)
        prior = _field_prior(term.get("field", ""))

        penalty = 0.0
        if term.get("obsolete"):
            penalty += 0.22
        if term.get("match_type") == "broad" or term.get("match_type") == "narrow":
            penalty += 0.06
        elif term.get("match_type") == "related":
            penalty += 0.12
        elif term.get("match_type") == "unclassified":
            penalty += 0.03

        score = (
            0.86 * lexical
            + 0.08 * harmonizer_score
            + 0.04 * source_score
            + 0.02 * prior
            - penalty
        )
        if lexical >= 0.98 and term.get("field") in {"primary_xref", "ncats_disease_id", "xref_id", "standard_name", "synonym"}:
            score += 0.02
        score = max(0.0, min(1.0, score))
        warnings: list[str] = []
        if term.get("obsolete"):
            warnings.append("obsolete_or_blocked_xref_match")
        if term.get("match_type") in {"broad", "narrow", "related"}:
            warnings.append(f"{term.get('match_type')}_xref_match")
        if source_count <= 1:
            warnings.append("single_source_concept")
        if score < 0.68:
            warnings.append("low_resolver_score")

        candidates.append({
            "primary_xref": pxref,
            "ncats_disease_id": concept.get("ncats_disease_id", ""),
            "standard_name": concept.get("standard_name", ""),
            "resolver_score": round(score, 4),
            "lexical_score": round(lexical, 4),
            "match_status": _resolver_match_status(score, lexical),
            "relationship_to_query": _relationship_to_query(term, lexical),
            "matched_term": term.get("term", ""),
            "matched_field": term.get("field", ""),
            "matched_source": term.get("source", ""),
            "matched_xref_id": term.get("xref_id", ""),
            "matched_xref_match_type": term.get("match_type", ""),
            "equivalence_confidence": term.get("equivalence_confidence", ""),
            "xref_confidence": term.get("xref_confidence", ""),
            "confidence_tier": concept.get("confidence_tier", ""),
            "overall_quality": concept.get("overall_quality", ""),
            "n_sources": concept.get("n_sources", ""),
            "is_rare": concept.get("is_rare", ""),
            "disease_type": concept.get("disease_type", ""),
            "hierarchy_parent_count": len(data.hierarchy_by_child.get(pxref, [])),
            "hierarchy_child_count": len(data.hierarchy_by_parent.get(pxref, [])),
            "warnings": warnings,
            "href": f"/disease-id-qa?ids={pxref}&tab=graph",
        })

    candidates.sort(key=lambda row: (-row["resolver_score"], -row["lexical_score"], row["standard_name"]))
    candidates = candidates[: max(1, min(limit, 50))]
    if len(candidates) > 1 and candidates[0]["resolver_score"] - candidates[1]["resolver_score"] <= 0.03:
        candidates[0]["warnings"] = sorted(set(candidates[0]["warnings"] + ["ambiguous_close_candidate"]))

    return {
        "query": query,
        "normalized_query": query_norm,
        "best": candidates[0] if candidates else None,
        "candidates": candidates,
    }


def bulk_resolve_names(
    data: DiseaseGraphData,
    queries: list[str],
    limit: int = 5,
    include_obsolete: bool = False,
) -> dict[str, Any]:
    clean_queries = [str(q).strip() for q in queries if str(q).strip()]
    return {
        "count": len(clean_queries),
        "limit": limit,
        "results": [
            resolve_name_candidates(
                data,
                query,
                limit=limit,
                include_obsolete=include_obsolete,
            )
            for query in clean_queries
        ],
    }


def _any_xref_matches(pxref: str, q_lower: str, data: DiseaseGraphData) -> bool:
    """Return True if any xref edge ID for this concept contains q_lower."""
    for edge in data.edges_by_pxref.get(pxref, []):
        if q_lower in edge.get("xref_id", "").lower():
            return True
    return False


def _match_filters(
    pxref: str,
    concept: dict,
    data: DiseaseGraphData,
    *,
    q_lower: str = "",
    filter_mode: str = "all",
    confidence_tier: str = "",
    is_rare: str = "",
    source: str = "",
    quality: str = "",
    disease_type: str = "",
) -> bool:
    """Return True if the concept passes all filter criteria."""
    # Flagged filter
    if filter_mode == "flagged":
        decisions = data.decisions_by_pxref.get(pxref, [])
        if not _is_flagged(concept, decisions):
            return False

    # Text search — checks primary_xref, name, synonyms, and xref IDs
    if q_lower:
        name = concept.get("standard_name", "").lower()
        synonyms = concept.get("synonyms", "").lower()
        if (q_lower not in pxref.lower()
                and q_lower not in name
                and q_lower not in synonyms
                and not _any_xref_matches(pxref, q_lower, data)):
            return False

    # Confidence tier
    if confidence_tier:
        if concept.get("confidence_tier", "").lower() != confidence_tier.lower():
            return False

    # Quality score — supports exact match or "min-max" range
    if quality:
        q_val = concept.get("overall_quality", "").strip()
        if "-" in quality:
            parts = quality.split("-", 1)
            try:
                q_min, q_max = float(parts[0]), float(parts[1])
                q_num = float(q_val) if q_val else None
                if q_num is None or not (q_min <= q_num < q_max):
                    return False
            except (ValueError, TypeError):
                return False
        else:
            if q_val != quality:
                return False

    # Rare disease
    if is_rare and is_rare != "all":
        concept_rare = concept.get("is_rare", "").lower() == "true"
        if is_rare == "true" and not concept_rare:
            return False
        if is_rare == "false" and concept_rare:
            return False

    # Source namespace
    if source:
        if not _concept_has_source(pxref, source, data):
            return False

    # Disease type (Biolink category)
    if disease_type:
        if concept.get("disease_type", "") != disease_type:
            return False

    return True


_SORT_KEYS: dict[str, str] = {
    "primary_xref": "primary_xref",
    "name": "standard_name",
    "tier": "confidence_tier",
    "quality": "overall_quality",
    "sources": "n_sources",
    "xrefs": "xref_count",
}


def search_concepts(
    data: DiseaseGraphData,
    q: str = "",
    filter_mode: str = "all",
    page: int = 1,
    per_page: int = 50,
    confidence_tier: str = "",
    is_rare: str = "",
    source: str = "",
    quality: str = "",
    disease_type: str = "",
    sort: str = "",
    sort_dir: str = "asc",
) -> dict[str, Any]:
    """Search concepts with pagination, optional filters, and sorting."""
    q_lower = q.strip().lower()
    matches: list[tuple[str, dict]] = []

    for pxref, concept in data.concepts_by_pxref.items():
        if not _match_filters(
            pxref, concept, data,
            q_lower=q_lower, filter_mode=filter_mode,
            confidence_tier=confidence_tier, is_rare=is_rare,
            source=source, quality=quality,
            disease_type=disease_type,
        ):
            continue
        matches.append((pxref, concept))

    # Sort if requested
    sort_field = _SORT_KEYS.get(sort, "")
    if sort_field:
        reverse = sort_dir.lower() == "desc"

        def _sort_val(pair: tuple[str, dict]):
            pxref, concept = pair
            if sort_field == "primary_xref":
                return pxref.lower()
            val = concept.get(sort_field, "")
            # Try numeric sort for quality/sources/xrefs
            try:
                return float(val)
            except (ValueError, TypeError):
                return str(val).lower()

        matches.sort(key=_sort_val, reverse=reverse)

    total = len(matches)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    page_items = matches[start : start + per_page]

    rows = [_concept_to_row(pxref, concept, data) for pxref, concept in page_items]

    return {
        "rows": rows,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "query": q,
        "filter": filter_mode,
    }


def export_search_tsv(
    data: DiseaseGraphData,
    q: str = "",
    filter_mode: str = "all",
    limit: int = 10_000,
) -> str:
    """Export search results as a TSV string (no pagination, capped at *limit* rows)."""
    q_lower = q.strip().lower()
    columns = [
        "primary_xref", "standard_name", "confidence_tier",
        "confidence_tier_original", "needs_review_auto_cleared",
        "needs_review_triage_bucket", "needs_review_triage_action",
        "needs_review_problem_namespaces", "needs_review_decision",
        "evidence_note", "overall_quality", "n_sources", "xref_count",
        "flags", "source_labels",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, delimiter="\t", extrasaction="ignore")
    writer.writeheader()

    count = 0
    for pxref, concept in data.concepts_by_pxref.items():
        if filter_mode == "flagged":
            decisions = data.decisions_by_pxref.get(pxref, [])
            if not _is_flagged(concept, decisions):
                continue
        if q_lower:
            name = concept.get("standard_name", "").lower()
            if q_lower not in pxref.lower() and q_lower not in name:
                continue
        writer.writerow(_concept_to_row(pxref, concept, data))
        count += 1
        if count >= limit:
            break

    return buf.getvalue()


_XREF_EDGE_COLUMNS = {
    "xref_id", "xref_namespace", "match_type", "relation_type",
    "match_type_source", "agreement_level", "xref_confidence",
    "relation_confidence", "equivalence_confidence",
    "mapping_confidence", "label_confidence", "source_support",
    "validation_support", "corroboration_score", "asserting_sources",
    "validation_sources", "evidence_streams", "structural_confidence",
    "evidence_tier", "review_tier", "score_rationale",
    "confidence_score", "label_similarity", "xref_label",
    "xref_is_obsolete", "source_asserted",
}

_CLINICAL_DESCENDANT_COLUMNS = {
    "anchor_xref_id", "anchor_xref_namespace", "anchor_xref_label",
    "anchor_match_type", "anchor_asserting_sources",
    "descendant_xref_id", "descendant_xref_namespace", "descendant_xref_label",
    "relationship_to_anchor", "relationship_to_disease", "predicate",
    "hierarchy_source", "distance_from_anchor", "distance_status", "path",
    "descendant_source_status", "existing_odin_match_types",
    "existing_odin_primary_xrefs", "in_mondo_xrefs", "in_yadaw_s2",
    "in_yadaw_s3", "rare_anchor", "may_not_be_rare_warning",
    "warning_reason", "evidence_sources", "review_status",
}

_CORE_DOWNLOAD_COLUMNS = ["ncats_disease_id", "primary_xref", "standard_name"]
_CONCEPT_DOWNLOAD_DEFAULT_COLUMNS = [
    *_CORE_DOWNLOAD_COLUMNS,
    "definition", "synonyms", "disease_type", "is_rare",
    "overall_quality", "n_sources", "xref_count",
    "nodenorm_canonical_curie", "nodenorm_canonical_label",
    "hierarchy_parents", "hierarchy_children",
    *_DISEASE_SOURCE_NAMESPACE_COLUMNS,
]
_XREF_EDGE_DOWNLOAD_DEFAULT_COLUMNS = [
    *_CORE_DOWNLOAD_COLUMNS,
    "xref_id", "xref_namespace", "xref_label", "match_type",
    "relation_type", "match_type_source", "equivalence_confidence",
    "relation_confidence", "evidence_tier", "review_tier",
    "xref_confidence", "agreement_level", "xref_is_obsolete",
    "asserting_sources", "validation_sources", "evidence_streams",
]
_CLINICAL_DESCENDANT_DOWNLOAD_DEFAULT_COLUMNS = [
    *_CORE_DOWNLOAD_COLUMNS,
    "disease_type", "is_rare",
    "anchor_xref_id", "anchor_xref_namespace", "anchor_xref_label",
    "anchor_match_type", "anchor_asserting_sources",
    "descendant_xref_id", "descendant_xref_namespace",
    "descendant_xref_label", "relationship_to_disease", "predicate",
    "hierarchy_source", "distance_from_anchor", "distance_status",
    "descendant_source_status", "may_not_be_rare_warning",
    "warning_reason", "review_status",
]


def _dedupe_columns(columns: list[str]) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for col in columns:
        if col and col not in seen:
            seen.add(col)
            deduped.append(col)
    return deduped


def _normalize_download_mode(mode: str = "concepts") -> str:
    mode = (mode or "concepts").strip().lower()
    aliases = {
        "concept": "concepts",
        "diseases": "concepts",
        "disease": "concepts",
        "edges": "xref_edges",
        "xref": "xref_edges",
        "xrefs": "xref_edges",
        "clinical": "clinical_descendants",
        "clinical_descendant": "clinical_descendants",
    }
    return aliases.get(mode, mode if mode in {"concepts", "xref_edges", "clinical_descendants"} else "concepts")


def _filtered_download_columns(
    columns: list[str] | None = None,
    mode: str = "concepts",
) -> tuple[list[str], bool, bool]:
    """Return (use_columns, xref_edge_mode, clinical_descendant_mode)."""
    mode = _normalize_download_mode(mode)
    if columns:
        use_columns = _dedupe_columns([*_CORE_DOWNLOAD_COLUMNS, *columns])
    elif mode == "xref_edges":
        use_columns = list(_XREF_EDGE_DOWNLOAD_DEFAULT_COLUMNS)
    elif mode == "clinical_descendants":
        use_columns = list(_CLINICAL_DESCENDANT_DOWNLOAD_DEFAULT_COLUMNS)
    else:
        use_columns = list(_CONCEPT_DOWNLOAD_DEFAULT_COLUMNS)
    if mode == "concepts":
        relationship_columns = _XREF_EDGE_COLUMNS | _CLINICAL_DESCENDANT_COLUMNS
        use_columns = [col for col in use_columns if col not in relationship_columns]
    elif mode == "xref_edges":
        use_columns = [col for col in use_columns if col not in _CLINICAL_DESCENDANT_COLUMNS]
    elif mode == "clinical_descendants":
        use_columns = [col for col in use_columns if col not in _XREF_EDGE_COLUMNS]
    clinical_descendant_mode = mode == "clinical_descendants"
    edge_mode = mode == "xref_edges"
    return use_columns, edge_mode, clinical_descendant_mode


def _filtered_download_rows(
    data: DiseaseGraphData,
    q: str = "",
    filter_mode: str = "all",
    confidence_tier: str = "",
    is_rare: str = "",
    source: str = "",
    quality: str = "",
    disease_type: str = "",
    columns: list[str] | None = None,
    mode: str = "concepts",
    limit: int = 200_000,
    include_sources: set[str] | None = None,
    exclude_obsolete: bool = False,
) -> tuple[list[str], bool, Any]:
    """Return (use_columns, edge_mode, row_iterator) for filtered downloads."""
    use_columns, edge_mode, clinical_descendant_mode = _filtered_download_columns(columns, mode)
    q_lower = q.strip().lower()

    def _iter_rows():
        count = 0
        for pxref, concept in data.concepts_by_pxref.items():
            if not _match_filters(
                pxref, concept, data,
                q_lower=q_lower, filter_mode=filter_mode,
                confidence_tier=confidence_tier, is_rare=is_rare,
                source=source, quality=quality,
                disease_type=disease_type,
            ):
                continue

            concept_row = _concept_to_row(
                pxref, concept, data,
                include_sources=include_sources,
                exclude_obsolete=exclude_obsolete,
            )

            if clinical_descendant_mode:
                descendant_rows = data.clinical_descendants_by_pxref.get(pxref, [])
                for descendant in descendant_rows:
                    ns = _normalize_ns(descendant.get("descendant_xref_namespace", ""))
                    anchor_ns = _normalize_ns(descendant.get("anchor_xref_namespace", ""))
                    if include_sources is not None and ns.upper() not in include_sources and anchor_ns.upper() not in include_sources:
                        continue
                    row = dict(concept_row)
                    for col in _CLINICAL_DESCENDANT_COLUMNS:
                        row[col] = descendant.get(col, "")
                    yield row
                    count += 1
                    if count >= limit:
                        return
            elif not edge_mode:
                yield concept_row
                count += 1
            else:
                edges = data.edges_by_pxref.get(pxref, [])
                for edge in edges:
                    ns = _normalize_ns(edge.get("xref_namespace", ""))
                    if include_sources is not None and ns.upper() not in include_sources:
                        continue
                    if exclude_obsolete and edge.get("xref_is_obsolete", "").lower() in ("true", "1"):
                        continue
                    row = dict(concept_row)
                    row["xref_id"] = edge.get("xref_id", "")
                    row["xref_namespace"] = ns
                    row["match_type"] = edge.get("match_type", "")
                    row["relation_type"] = edge.get("relation_type", "") or edge.get("match_type", "")
                    row["match_type_source"] = edge.get("match_type_source", "")
                    row["agreement_level"] = edge.get("agreement_level", "")
                    row["xref_confidence"] = edge.get("xref_confidence", "")
                    row["relation_confidence"] = edge.get("relation_confidence", "")
                    row["equivalence_confidence"] = edge.get("equivalence_confidence", "")
                    row["mapping_confidence"] = edge.get("mapping_confidence", "")
                    row["label_confidence"] = edge.get("label_confidence", "")
                    row["source_support"] = edge.get("source_support", "")
                    row["validation_support"] = edge.get("validation_support", "")
                    row["corroboration_score"] = edge.get("corroboration_score", "")
                    row["asserting_sources"] = edge.get("asserting_sources", "")
                    row["validation_sources"] = edge.get("validation_sources", "")
                    row["evidence_streams"] = edge.get("evidence_streams", "")
                    row["structural_confidence"] = edge.get("structural_confidence", "")
                    row["evidence_tier"] = edge.get("evidence_tier", "")
                    row["review_tier"] = edge.get("review_tier", "")
                    row["score_rationale"] = edge.get("score_rationale", "")
                    row["confidence_score"] = edge.get("confidence_score", "")
                    row["label_similarity"] = edge.get("label_similarity", "")
                    row["xref_label"] = edge.get("xref_label", "")
                    row["xref_is_obsolete"] = edge.get("xref_is_obsolete", "")
                    row["source_asserted"] = edge.get("source_asserted", "")
                    yield row
                    count += 1
                    if count >= limit:
                        return

            if count >= limit:
                return

    return use_columns, edge_mode, _iter_rows()


def compute_download_preview_counts(
    data: DiseaseGraphData,
    q: str = "",
    filter_mode: str = "all",
    confidence_tier: str = "",
    is_rare: str = "",
    source: str = "",
    quality: str = "",
    disease_type: str = "",
    mode: str = "concepts",
    include_sources: set[str] | None = None,
    exclude_obsolete: bool = False,
) -> dict[str, Any]:
    """Return exact concept and row counts for the selected download mode."""
    mode_key = _normalize_download_mode(mode)
    q_lower = q.strip().lower()
    concept_count = 0
    xref_edge_count = 0
    clinical_descendant_count = 0

    for pxref, concept in data.concepts_by_pxref.items():
        if not _match_filters(
            pxref, concept, data,
            q_lower=q_lower, filter_mode=filter_mode,
            confidence_tier=confidence_tier, is_rare=is_rare,
            source=source, quality=quality,
            disease_type=disease_type,
        ):
            continue
        concept_count += 1

        for edge in data.edges_by_pxref.get(pxref, []):
            ns = _normalize_ns(edge.get("xref_namespace", ""))
            if include_sources is not None and ns.upper() not in include_sources:
                continue
            if exclude_obsolete and edge.get("xref_is_obsolete", "").lower() in ("true", "1"):
                continue
            xref_edge_count += 1

        for descendant in data.clinical_descendants_by_pxref.get(pxref, []):
            ns = _normalize_ns(descendant.get("descendant_xref_namespace", ""))
            anchor_ns = _normalize_ns(descendant.get("anchor_xref_namespace", ""))
            if include_sources is not None and ns.upper() not in include_sources and anchor_ns.upper() not in include_sources:
                continue
            clinical_descendant_count += 1

    row_count = {
        "xref_edges": xref_edge_count,
        "clinical_descendants": clinical_descendant_count,
    }.get(mode_key, concept_count)
    row_label = {
        "xref_edges": "xref edge rows",
        "clinical_descendants": "clinical descendant rows",
    }.get(mode_key, "concept rows")

    return {
        "mode": mode_key,
        "row_count": row_count,
        "row_label": row_label,
        "concept_count": concept_count,
        "xref_edge_count": xref_edge_count,
        "clinical_descendant_count": clinical_descendant_count,
    }


def export_filtered_download(
    data: DiseaseGraphData,
    q: str = "",
    filter_mode: str = "all",
    confidence_tier: str = "",
    is_rare: str = "",
    source: str = "",
    quality: str = "",
    disease_type: str = "",
    columns: list[str] | None = None,
    mode: str = "concepts",
    fmt: str = "tsv",
    limit: int = 200_000,
    include_sources: set[str] | None = None,
    exclude_obsolete: bool = False,
) -> str:
    """Export filtered concepts with selectable columns.

    When any xref edge column is selected, the output switches from one row
    per concept to one row per xref edge (concept fields repeated on each
    row).

    Parameters
    ----------
    columns : list[str] | None
        Column names to include. ``None`` uses the default set.
    fmt : str
        "tsv" or "csv".
    include_sources : set[str] | None
        When set, only include xref edges whose namespace (upper-cased)
        is in this set.  ``None`` means include all sources.
    exclude_obsolete : bool
        When True, omit xref edges where ``xref_is_obsolete`` is true.
    """
    use_columns, edge_mode, rows = _filtered_download_rows(
        data, q=q, filter_mode=filter_mode,
        confidence_tier=confidence_tier, is_rare=is_rare, source=source,
        quality=quality, disease_type=disease_type, columns=columns,
        mode=mode,
        limit=limit, include_sources=include_sources,
        exclude_obsolete=exclude_obsolete,
    )
    delimiter = "," if fmt == "csv" else "\t"
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=use_columns, delimiter=delimiter, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


def stream_filtered_download(
    data: DiseaseGraphData,
    q: str = "",
    filter_mode: str = "all",
    confidence_tier: str = "",
    is_rare: str = "",
    source: str = "",
    quality: str = "",
    disease_type: str = "",
    columns: list[str] | None = None,
    mode: str = "concepts",
    fmt: str = "tsv",
    limit: int = 200_000,
    include_sources: set[str] | None = None,
    exclude_obsolete: bool = False,
):
    """Yield lines for streaming TSV/CSV download."""
    use_columns, edge_mode, rows = _filtered_download_rows(
        data, q=q, filter_mode=filter_mode,
        confidence_tier=confidence_tier, is_rare=is_rare, source=source,
        quality=quality, disease_type=disease_type, columns=columns,
        mode=mode,
        limit=limit, include_sources=include_sources,
        exclude_obsolete=exclude_obsolete,
    )
    delimiter = "," if fmt == "csv" else "\t"

    # Header line
    line_buf = io.StringIO()
    writer = csv.DictWriter(line_buf, fieldnames=use_columns, delimiter=delimiter, extrasaction="ignore")
    writer.writeheader()
    yield line_buf.getvalue()

    # Data rows
    for row in rows:
        line_buf = io.StringIO()
        writer = csv.DictWriter(line_buf, fieldnames=use_columns, delimiter=delimiter, extrasaction="ignore")
        writer.writerow(row)
        yield line_buf.getvalue()


def export_filtered_xlsx(
    data: DiseaseGraphData,
    q: str = "",
    filter_mode: str = "all",
    confidence_tier: str = "",
    is_rare: str = "",
    source: str = "",
    quality: str = "",
    disease_type: str = "",
    columns: list[str] | None = None,
    mode: str = "concepts",
    limit: int = 200_000,
    include_sources: set[str] | None = None,
    exclude_obsolete: bool = False,
) -> bytes:
    """Export filtered concepts as an Excel (.xlsx) file returned as bytes."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    use_columns, edge_mode, rows = _filtered_download_rows(
        data, q=q, filter_mode=filter_mode,
        confidence_tier=confidence_tier, is_rare=is_rare, source=source,
        quality=quality, disease_type=disease_type, columns=columns,
        mode=mode,
        limit=limit, include_sources=include_sources,
        exclude_obsolete=exclude_obsolete,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = {
        "xref_edges": "Disease Xref Edges",
        "clinical_descendants": "Clinical Descendants",
    }.get(_normalize_download_mode(mode), "Disease Concepts")

    # Header
    ws.append(use_columns)

    # Data
    for row in rows:
        ws.append([row.get(col, "") for col in use_columns])

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Auto-width columns (capped at 40)
    for col_idx, col_name in enumerate(use_columns, 1):
        max_len = len(col_name)
        for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=min(100, ws.max_row)):
            for cell in row_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# SSSOM Export
# ---------------------------------------------------------------------------

_MATCH_TYPE_TO_SKOS: dict[str, str] = {
    "exact": "skos:exactMatch",
    "broad": "skos:broadMatch",
    "narrow": "skos:narrowMatch",
    "related": "skos:relatedMatch",
}

_MATCH_SOURCE_TO_SEMAPV: dict[str, str] = {
    "lexical": "semapv:LexicalMatching",
    "logical": "semapv:LogicalReasoning",
    "manual": "semapv:ManualMappingCuration",
    "composite": "semapv:CompositeMatching",
    "semantic": "semapv:SemanticSimilarityThresholdMatching",
}

_CURIE_MAP: dict[str, str] = {
    "MONDO": "http://purl.obolibrary.org/obo/MONDO_",
    "OMIM": "https://omim.org/entry/",
    "DOID": "http://purl.obolibrary.org/obo/DOID_",
    "Orphanet": "http://www.orpha.net/ORDO/Orphanet_",
    "MedGen": "https://www.ncbi.nlm.nih.gov/medgen/",
    "GARD": "https://rarediseases.info.nih.gov/diseases/",
    "MESH": "http://id.nlm.nih.gov/mesh/",
    "UMLS": "https://uts.nlm.nih.gov/uts/umls/concept/",
    "EFO": "http://www.ebi.ac.uk/efo/EFO_",
    "HP": "http://purl.obolibrary.org/obo/HP_",
    "skos": "http://www.w3.org/2004/02/skos/core#",
    "semapv": "https://w3id.org/semapv/vocab/",
}


def _match_type_to_skos(match_type: str) -> str:
    """Map a match_type string to the corresponding SKOS predicate CURIE."""
    return _MATCH_TYPE_TO_SKOS.get(match_type, "skos:closeMatch")


def _match_source_to_semapv(match_type_source: str) -> str:
    """Map a match_type_source string to a semapv justification CURIE."""
    src = match_type_source.strip().lower()
    if src == "gard_curated_xref":
        return "semapv:ManualMappingCuration"
    for key, val in _MATCH_SOURCE_TO_SEMAPV.items():
        if key in src:
            return val
    return "semapv:UnspecifiedMatching"


def _iter_sssom_lines(
    data: DiseaseGraphData,
    include_sources: set[str] | None = None,
):
    """Yield SSSOM TSV lines one at a time."""
    version = data.manifest.get("pipeline_version", "unknown")

    # Metadata header
    yield "#curie_map:"
    for prefix, uri in _CURIE_MAP.items():
        yield f"#  {prefix}: {uri}"
    yield f"#mapping_set_id: https://github.com/ncats/IFX_ODIN/disease-harmonizer"
    yield f"#mapping_set_version: {version}"
    yield "#creator_label: TargetGraph Disease Harmonizer"
    yield ""

    # TSV header
    columns = [
        "subject_id", "subject_label", "predicate_id",
        "object_id", "object_label",
        "mapping_justification", "confidence",
        "subject_source", "object_source", "comment",
    ]
    yield "\t".join(columns)

    # Pre-compute upper-cased source filter set once
    upper_sources = {s.upper() for s in include_sources} if include_sources else None

    for pxref, concept in data.concepts_by_pxref.items():
        for edge in data.edges_by_pxref.get(pxref, []):
            ns = _normalize_ns(edge.get("xref_namespace", ""))
            if upper_sources and ns.upper() not in upper_sources:
                continue
            xref_id = edge.get("xref_id", "")
            if not xref_id:
                continue
            match_type = edge.get("match_type", "unclassified")
            match_source = edge.get("match_type_source", "")
            label_row = data.labels_by_xref_id.get(xref_id, {})
            object_label = edge.get("xref_label", "") or label_row.get("preferred_label", "")
            if object_label in ("nan", "NaN"):
                object_label = ""
            confidence = edge.get("equivalence_confidence", "") or edge.get("xref_confidence", "")
            row = [
                pxref,
                concept.get("standard_name", ""),
                _match_type_to_skos(match_type),
                xref_id,
                object_label,
                _match_source_to_semapv(match_source),
                confidence,
                concept.get("primary_xref_source", ""),
                ns,
                edge.get("score_rationale", "") or edge.get("evidence_tier", "") or edge.get("agreement_level", ""),
            ]
            yield "\t".join(row)


def export_sssom(
    data: DiseaseGraphData,
    include_sources: set[str] | None = None,
) -> str:
    """Export xref edges as SSSOM TSV (in-memory, for backward compat)."""
    return "\n".join(_iter_sssom_lines(data, include_sources)) + "\n"


def iter_sssom_bytes(
    data: DiseaseGraphData,
    include_sources: set[str] | None = None,
):
    """Yield SSSOM TSV as encoded byte chunks for streaming."""
    for line in _iter_sssom_lines(data, include_sources):
        yield (line + "\n").encode("utf-8")


# ---------------------------------------------------------------------------
# Public REST API helpers
# ---------------------------------------------------------------------------

def resolve_concept(data: DiseaseGraphData, curie: str) -> dict[str, Any] | None:
    """Resolve a CURIE to full concept data including all xref edges and decisions."""
    pxref = _resolve_to_pxref(data, curie)
    if pxref is None:
        return None
    concept = data.concepts_by_pxref[pxref]
    edges = data.edges_by_pxref.get(pxref, [])
    decisions = data.decisions_by_pxref.get(pxref, [])

    xrefs = []
    for edge in edges:
        xrefs.append({
            "xref_id": edge.get("xref_id", ""),
            "namespace": _normalize_ns(edge.get("xref_namespace", "")),
            "match_type": edge.get("match_type", ""),
            "relation_type": edge.get("relation_type", "") or edge.get("match_type", ""),
            "confidence": edge.get("equivalence_confidence", "") or edge.get("xref_confidence", ""),
            "equivalence_confidence": edge.get("equivalence_confidence", ""),
            "relation_confidence": edge.get("relation_confidence", ""),
            "evidence_tier": edge.get("evidence_tier", ""),
            "review_tier": edge.get("review_tier", ""),
            "agreement_level": edge.get("agreement_level", ""),
            "match_type_source": edge.get("match_type_source", ""),
            "xref_label": edge.get("xref_label", ""),
            "xref_is_obsolete": edge.get("xref_is_obsolete", ""),
            "mapping_confidence": edge.get("mapping_confidence", ""),
            "label_confidence": edge.get("label_confidence", ""),
            "source_support": edge.get("source_support", ""),
            "validation_support": edge.get("validation_support", ""),
            "corroboration_score": edge.get("corroboration_score", ""),
            "asserting_sources": edge.get("asserting_sources", ""),
            "validation_sources": edge.get("validation_sources", ""),
            "evidence_streams": edge.get("evidence_streams", ""),
            "structural_confidence": edge.get("structural_confidence", ""),
            "score_rationale": edge.get("score_rationale", ""),
        })

    decision_list = []
    for dec in decisions:
        decision_list.append({
            "decision_id": dec.get("decision_id", ""),
            "decision_type": dec.get("decision_type", ""),
            "xref_id": dec.get("xref_id", ""),
            "status": dec.get("status", ""),
            "auto_decision": dec.get("auto_decision", ""),
            "auto_confidence": dec.get("auto_confidence", ""),
            "auto_rationale": dec.get("auto_rationale", ""),
        })

    parents = [
        {
            "primary_xref": row.get("parent_primary_xref", ""),
            "ncats_disease_id": row.get("parent_ncats_disease_id", ""),
            "label": row.get("parent_label", ""),
            "predicate": row.get("predicate", ""),
            "source": row.get("source", ""),
        }
        for row in data.hierarchy_by_child.get(pxref, [])
    ]
    children = [
        {
            "primary_xref": row.get("child_primary_xref", ""),
            "ncats_disease_id": row.get("child_ncats_disease_id", ""),
            "label": row.get("child_label", ""),
            "predicate": row.get("predicate", ""),
            "source": row.get("source", ""),
        }
        for row in data.hierarchy_by_parent.get(pxref, [])
    ]

    return {
        "ncats_disease_id": concept.get("ncats_disease_id", ""),
        "primary_xref": pxref,
        "standard_name": concept.get("standard_name", ""),
        "confidence_tier": concept.get("confidence_tier", ""),
        "overall_quality": concept.get("overall_quality", ""),
        "n_sources": concept.get("n_sources", ""),
        "disease_type": concept.get("disease_type", ""),
        "is_rare": concept.get("is_rare", ""),
        "definition": concept.get("definition", ""),
        "synonyms": concept.get("synonyms", ""),
        "xrefs": xrefs,
        "decisions": decision_list,
        "hierarchy": {
            "parents": parents,
            "children": children,
        },
        "nodenorm": {
            "canonical_curie": concept.get("nodenorm_canonical_curie", ""),
            "canonical_label": concept.get("nodenorm_canonical_label", ""),
            "validation_status": _nodenorm_status(concept),
        },
    }


# ---------------------------------------------------------------------------
# Source Agreement Matrix (for heatmap)
# ---------------------------------------------------------------------------

_NS_CASE_NORMALIZE: dict[str, str] = {
    "doid": "DOID",
    "efo": "EFO",
    "gard": "GARD",
    "hp": "HP",
    "icd10": "ICD10",
    "icd10cm": "ICD10CM",
    "icd11": "ICD11",
    "icd11f": "ICD11f",
    "icd9": "ICD9",
    "icdo": "ICDO",
    "kegg": "KEGG",
    "meddra": "MEDDRA",
    "snomedct": "SNOMEDCT",
    "medgen": "MEDGEN",
    "mesh": "MESH",
    "mondo": "MONDO",
    "ncit": "NCIT",
    "omim": "OMIM",
    "omimps": "OMIMPS",
    "ordo": "ORDO",
    "orphanet": "Orphanet",
    "umls": "UMLS",
}


def _normalize_ns(ns: str) -> str:
    """Normalize namespace casing (e.g. snomedct → SNOMEDCT)."""
    ns = (ns or "").strip()
    return _NS_CASE_NORMALIZE.get(ns.lower(), ns)


def compute_source_agreement_matrix(data: DiseaseGraphData) -> dict[str, Any]:
    """Compute pairwise source co-occurrence overlap for the heatmap.

    For each pair of sources (A, B) this measures Jaccard overlap:
    how many concepts have xrefs from both A and B, divided by concepts
    that have xrefs from either A or B.  This shows which sources tend
    to cover the same disease space.
    """
    # Build per-source concept sets (with namespace normalization)
    source_concepts: dict[str, set[str]] = defaultdict(set)
    for pxref, edges in data.edges_by_pxref.items():
        seen: set[str] = set()
        for edge in edges:
            ns = _normalize_ns(edge.get("xref_namespace", "").strip())
            if ns and ns not in seen:
                seen.add(ns)
                source_concepts[ns].add(pxref)

    # Pick top sources by concept count, exclude very small ones
    ranked = sorted(source_concepts.keys(), key=lambda k: -len(source_concepts[k]))
    sources = [s for s in ranked if len(source_concepts[s]) >= 500][:10]
    if not sources:
        sources = ranked[:6]

    n = len(sources)
    overlap_counts = [[0] * n for _ in range(n)]
    union_counts = [[0] * n for _ in range(n)]
    matrix = [[0.0] * n for _ in range(n)]

    for i in range(n):
        for j in range(i, n):
            sa, sb = source_concepts[sources[i]], source_concepts[sources[j]]
            ov = len(sa & sb)
            un = len(sa | sb)
            overlap_counts[i][j] = ov
            overlap_counts[j][i] = ov
            union_counts[i][j] = un
            union_counts[j][i] = un
            rate = round(ov / un, 4) if un > 0 else 0.0
            matrix[i][j] = rate
            matrix[j][i] = rate

    return {
        "sources": sources,
        "matrix": matrix,
        "counts": overlap_counts,
        "union_counts": union_counts,
    }


# ---------------------------------------------------------------------------
# Source Flow Data (for Sankey diagram)
# ---------------------------------------------------------------------------

def compute_source_flow_data(data: DiseaseGraphData) -> dict[str, Any]:
    """Compute flows: source namespace -> confidence tier."""
    # Count: for each (namespace, tier), how many concepts have an edge in that namespace
    flow_counts: Counter[tuple[str, str]] = Counter()
    all_ns: set[str] = set()
    all_tiers: set[str] = set()

    for pxref, concept in data.concepts_by_pxref.items():
        tier = concept.get("confidence_tier", "").strip() or "unknown"
        all_tiers.add(tier)
        seen_ns: set[str] = set()
        for edge in data.edges_by_pxref.get(pxref, []):
            ns = _normalize_ns(edge.get("xref_namespace", "").strip())
            if ns and ns not in seen_ns:
                seen_ns.add(ns)
                all_ns.add(ns)
                flow_counts[(ns, tier)] += 1

    # Build node list — only top namespaces by count
    ns_totals = Counter({ns: sum(flow_counts[(ns, t)] for t in all_tiers) for ns in all_ns})
    top_ns = [ns for ns, _ in ns_totals.most_common(8)]

    nodes: list[dict[str, str]] = []
    ns_ids: dict[str, str] = {}
    tier_ids: dict[str, str] = {}

    for ns in top_ns:
        nid = f"src_{ns}"
        ns_ids[ns] = nid
        nodes.append({"id": nid, "label": ns, "type": "source"})

    tier_order = [
        "multi_source_supported",
        "single_authoritative_source",
        "needs_review",
        "review_validator_disagreement",
        "review_obsolete_xref",
    ]
    for tier in tier_order:
        if tier in all_tiers:
            tid = f"tier_{tier}"
            tier_ids[tier] = tid
            nodes.append({"id": tid, "label": tier, "type": "tier"})
    # Add any remaining tiers
    for tier in sorted(all_tiers):
        if tier not in tier_ids:
            tid = f"tier_{tier}"
            tier_ids[tier] = tid
            nodes.append({"id": tid, "label": tier, "type": "tier"})

    links: list[dict[str, Any]] = []
    for ns in top_ns:
        for tier in all_tiers:
            count = flow_counts.get((ns, tier), 0)
            if count > 0 and ns in ns_ids and tier in tier_ids:
                links.append({
                    "source": ns_ids[ns],
                    "target": tier_ids[tier],
                    "value": count,
                })

    return {"nodes": nodes, "links": links}


# ---------------------------------------------------------------------------
# Version Diff
# ---------------------------------------------------------------------------

_VERSION_CACHE_MAX = 4
_version_graph_cache: OrderedDict[str, DiseaseGraphData] = OrderedDict()


def load_version_data(data_dir: str | Path) -> DiseaseGraphData:
    """Load any versioned app_graph directory, cached by absolute path."""
    data_dir = Path(data_dir)
    cache_key = str(data_dir.resolve())
    if cache_key in _version_graph_cache:
        _version_graph_cache.move_to_end(cache_key)
        return _version_graph_cache[cache_key]
    data = _load_app_graph_data(data_dir)
    _version_graph_cache[cache_key] = data
    while len(_version_graph_cache) > _VERSION_CACHE_MAX:
        _version_graph_cache.popitem(last=False)
    _print_graph_load(f"Disease version {data_dir.name}", data)
    return data


def load_baseline_data(baseline_dir: str | Path) -> DiseaseGraphData:
    """Backward-compatible wrapper for version comparison."""
    return load_version_data(baseline_dir)


def _concept_biolink_category(concept: dict) -> str:
    return concept.get("disease_type", "").strip() or "unknown"


def _concept_type_distribution(data: DiseaseGraphData) -> Counter[str]:
    return Counter(_concept_biolink_category(concept) for concept in data.concepts_by_pxref.values())


def _edge_namespace_distribution(data: DiseaseGraphData) -> Counter[str]:
    counts: Counter[str] = Counter()
    for edges in data.edges_by_pxref.values():
        for edge in edges:
            ns = _normalize_ns(edge.get("xref_namespace", "").strip()) or "unknown"
            counts[ns] += 1
    return counts


def _xref_owner_index(data: DiseaseGraphData) -> dict[str, set[str]]:
    owners: dict[str, set[str]] = defaultdict(set)
    for pxref, edges in data.edges_by_pxref.items():
        for edge in edges:
            xid = edge.get("xref_id", "")
            if xid:
                owners[xid].add(pxref)
    return owners


def _distribution_delta(
    baseline: Counter[str],
    current: Counter[str],
    key_name: str,
) -> list[dict[str, Any]]:
    rows = []
    for key in sorted(set(baseline) | set(current)):
        old_count = int(baseline.get(key, 0))
        new_count = int(current.get(key, 0))
        rows.append({
            key_name: key,
            "old_count": old_count,
            "new_count": new_count,
            "delta": new_count - old_count,
        })
    return sorted(rows, key=lambda r: (-abs(r["delta"]), r[key_name]))


def _clean_source_version_value(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"", "unknown", "none", "nan", "null"} else text


def _disease_source_version_map(data: DiseaseGraphData) -> dict[str, dict[str, str]]:
    rows = data.source_catalog or data.manifest.get("source_versions") or []
    out: dict[str, dict[str, str]] = {}
    if not isinstance(rows, list):
        return out
    for raw in rows:
        if not isinstance(raw, dict):
            continue
        source = _clean_source_version_value(
            raw.get("source_name") or raw.get("source") or raw.get("name")
        )
        if not source:
            continue
        out[source] = {
            "source": source,
            "version": _clean_source_version_value(raw.get("source_version") or raw.get("version")),
            "release_date": _clean_source_version_value(raw.get("release_date") or raw.get("source_release_date")),
            "download_start": _clean_source_version_value(raw.get("download_start") or raw.get("odin_download_date")),
            "metadata_file": _clean_source_version_value(raw.get("metadata_file")),
        }
    return out


def _disease_source_version_changes(
    baseline: DiseaseGraphData,
    current: DiseaseGraphData,
) -> list[dict[str, Any]]:
    baseline_versions = _disease_source_version_map(baseline)
    current_versions = _disease_source_version_map(current)
    ordered_sources: list[str] = []
    for source in list(current_versions) + list(baseline_versions):
        if source not in ordered_sources:
            ordered_sources.append(source)

    rows: list[dict[str, Any]] = []
    for source in ordered_sources:
        old = baseline_versions.get(source, {})
        new = current_versions.get(source, {})
        old_version = old.get("version", "")
        new_version = new.get("version", "")
        if old and not new:
            status = "removed"
        elif new and not old:
            status = "added"
        elif old_version != new_version:
            status = "version_changed"
        else:
            status = "unchanged"
        rows.append({
            "source": source,
            "old_version": old_version,
            "new_version": new_version,
            "status": status,
            "old_download_date": (old.get("download_start", "") or old.get("release_date", ""))[:10],
            "new_download_date": (new.get("download_start", "") or new.get("release_date", ""))[:10],
        })
    return rows


def _format_signed_delta(value: int, label: str) -> str:
    sign = "+" if value > 0 else ""
    return f"{label} {sign}{value:,}"


def _disease_change_drivers(
    source_version_rows: list[dict[str, Any]],
    category_delta: list[dict[str, Any]],
    namespace_delta: list[dict[str, Any]],
    data_layer_changes: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    source_rows = {row["source"]: row for row in source_version_rows}
    category_counts = {row["category"]: int(row.get("delta") or 0) for row in category_delta}
    ns_counts = {row["namespace"]: int(row.get("delta") or 0) for row in namespace_delta}
    layer_counts = {row["layer"]: int(row.get("delta") or 0) for row in data_layer_changes}

    def version_text(source: str) -> str:
        row = source_rows.get(source, {})
        old = row.get("old_version") or "not captured"
        new = row.get("new_version") or "not captured"
        return f"{old} -> {new}"

    nodenorm_changed = source_rows.get("NodeNorm", {}).get("status") != "unchanged"
    nodenorm_interpretation = (
        "NodeNorm advanced in this release; this affects preferred CURIE validation and label/canonicalization context, not primary source authority."
        if nodenorm_changed
        else "NodeNorm stayed version-stable; it should be interpreted as validator context rather than an asserting xref source."
    )

    return [
        {
            "source": "NodeNorm",
            "version_change": version_text("NodeNorm"),
            "observed_delta": "; ".join([
                _format_signed_delta(ns_counts.get("UMLS", 0), "UMLS xref edges"),
                _format_signed_delta(layer_counts.get("Xref labels", 0), "xref label rows"),
            ]),
            "interpretation": nodenorm_interpretation,
        },
        {
            "source": "SNOMEDCT / ICD-10-CM",
            "version_change": f"SNOMEDCT {version_text('SNOMEDCT')}; ICD-10-CM {version_text('ICD-10-CM')}",
            "observed_delta": "; ".join([
                _format_signed_delta(ns_counts.get("SNOMEDCT", 0), "SNOMEDCT xref edges"),
                _format_signed_delta(ns_counts.get("ICD10", 0) + ns_counts.get("ICD10CM", 0), "ICD-10 xref edges"),
                _format_signed_delta(layer_counts.get("Clinical descendant edges", 0), "clinical descendant edges"),
            ]),
            "interpretation": "Clinical-code movement should be interpreted separately from exact disease xrefs, especially when descendant-only SNOMEDCT/ICD codes are included as context.",
        },
        {
            "source": "MONDO / DOID / EFO",
            "version_change": f"MONDO {version_text('MONDO')}; DOID {version_text('DOID')}; EFO {version_text('EFO')}",
            "observed_delta": "; ".join([
                _format_signed_delta(ns_counts.get("MONDO", 0), "MONDO xref edges"),
                _format_signed_delta(ns_counts.get("DOID", 0), "DOID xref edges"),
                _format_signed_delta(ns_counts.get("EFO", 0), "EFO xref edges"),
                _format_signed_delta(layer_counts.get("Ontology hierarchy edges", 0), "hierarchy edges"),
            ]),
            "interpretation": "Ontology-source changes drive concept grounding, hierarchy context, and curated exact/broad/narrow match evidence.",
        },
        {
            "source": "MedGen / OMIM / Orphanet",
            "version_change": f"MedGen {version_text('MedGen')}; OMIM {version_text('OMIM')}; Orphanet {version_text('Orphanet')}",
            "observed_delta": "; ".join([
                _format_signed_delta(ns_counts.get("MEDGEN", 0), "MedGen xref edges"),
                _format_signed_delta(ns_counts.get("OMIM", 0) + ns_counts.get("OMIMPS", 0), "OMIM xref edges"),
                _format_signed_delta(ns_counts.get("Orphanet", 0), "Orphanet xref edges"),
            ]),
            "interpretation": "Disease-source refreshes change the cross-reference evidence used by ODIN; downstream QC determines whether conflicts remain open or are resolved.",
        },
        {
            "source": "Gene association sources",
            "version_change": f"ClinGen {version_text('ClinGen')}; Monarch {version_text('Monarch')}; Jensen {version_text('Jensen')}; MedGen associations {version_text('MedGen Gene Associations')}",
            "observed_delta": "; ".join([
                _format_signed_delta(layer_counts.get("Disease-gene associations", 0), "disease-gene associations"),
                _format_signed_delta(category_counts.get("biolink:PhenotypicFeature", 0), "phenotype-labeled concepts"),
            ]),
            "interpretation": "Association-source changes affect the disease-gene context layer and should not be confused with disease xref equivalence changes.",
        },
    ]


def _xref_edge_key(pxref: str, edge: dict) -> tuple[str, str, str]:
    return (
        pxref,
        edge.get("xref_id", ""),
        _normalize_ns(edge.get("xref_namespace", "").strip()),
    )


def compute_version_diff(
    current: DiseaseGraphData,
    baseline: DiseaseGraphData,
) -> dict[str, Any]:
    """Compute delta between two versioned datasets."""
    current_pxrefs = set(current.concepts_by_pxref.keys())
    baseline_pxrefs = set(baseline.concepts_by_pxref.keys())

    # Build ncats_disease_id → primary_xref maps for stable identity matching
    baseline_ncats_to_pxref: dict[str, str] = {}
    for pxref, concept in baseline.concepts_by_pxref.items():
        nid = concept.get("ncats_disease_id", "").strip()
        if nid:
            baseline_ncats_to_pxref[nid] = pxref

    current_ncats_to_pxref: dict[str, str] = {}
    for pxref, concept in current.concepts_by_pxref.items():
        nid = concept.get("ncats_disease_id", "").strip()
        if nid:
            current_ncats_to_pxref[nid] = pxref

    shared_ncats_ids = set(baseline_ncats_to_pxref) & set(current_ncats_to_pxref)

    # Concepts matched by stable ncats_disease_id (even if primary_xref changed)
    matched_baseline_pxrefs: set[str] = set()
    matched_current_pxrefs: set[str] = set()
    identity_matched_pairs: list[tuple[str, str]] = []  # (baseline_pxref, current_pxref)
    rekeyed_concepts: list[dict[str, str]] = []

    for nid in shared_ncats_ids:
        b_pxref = baseline_ncats_to_pxref[nid]
        c_pxref = current_ncats_to_pxref[nid]
        matched_baseline_pxrefs.add(b_pxref)
        matched_current_pxrefs.add(c_pxref)
        identity_matched_pairs.append((b_pxref, c_pxref))
        if b_pxref != c_pxref:
            rekeyed_concepts.append({
                "ncats_disease_id": nid,
                "old_primary_xref": b_pxref,
                "new_primary_xref": c_pxref,
                "standard_name": current.concepts_by_pxref[c_pxref].get("standard_name", ""),
            })

    current_xref_owners = _xref_owner_index(current)
    baseline_xref_owners = _xref_owner_index(baseline)

    # Raw primary-row churn can be large when the preferred primary grounding
    # changes between releases. Treat previous primary IDs that are now current
    # xrefs as grounding changes before counting true concept additions/removals.
    raw_added_pxrefs = current_pxrefs - matched_current_pxrefs - baseline_pxrefs
    raw_removed_pxrefs = baseline_pxrefs - matched_baseline_pxrefs - current_pxrefs
    primary_grounding_changes: list[dict[str, Any]] = []
    previous_primary_represented_as_xref: set[str] = set()
    current_primary_explained_by_previous_xref: set[str] = set()
    for old_pxref in sorted(raw_removed_pxrefs):
        owners = sorted(current_xref_owners.get(old_pxref, set()) - {old_pxref})
        if not owners:
            continue
        previous_primary_represented_as_xref.add(old_pxref)
        current_primary_explained_by_previous_xref.update(owners)
        owner_labels = [
            current.concepts_by_pxref[owner].get("standard_name", "")
            for owner in owners[:3]
            if owner in current.concepts_by_pxref
        ]
        owner_suffix = "" if len(owners) <= 3 else f" +{len(owners) - 3} more"
        primary_grounding_changes.append({
            "old_primary_xref": old_pxref,
            "old_standard_name": baseline.concepts_by_pxref[old_pxref].get("standard_name", ""),
            "new_primary_xref": ", ".join(owners[:3]) + owner_suffix,
            "new_standard_name": " | ".join(owner_labels),
            "n_current_primary_concepts": len(owners),
            "reason": "Previous primary identifier is now represented as a current xref",
        })

    # Concept additions/removals after primary-grounding changes are accounted for.
    added_pxrefs = raw_added_pxrefs - current_primary_explained_by_previous_xref
    removed_pxrefs = raw_removed_pxrefs - previous_primary_represented_as_xref

    # Concepts matched by pxref that weren't already matched by ncats_id
    shared_pxrefs = (current_pxrefs & baseline_pxrefs) - matched_current_pxrefs

    # Tier / quality / category changes across identity-matched pairs
    tier_changes: list[dict[str, str]] = []
    quality_changes: list[dict[str, str]] = []
    disease_type_changes: list[dict[str, str]] = []

    def _compare_pair(b_pxref: str, c_pxref: str) -> None:
        cur = current.concepts_by_pxref[c_pxref]
        base = baseline.concepts_by_pxref[b_pxref]
        cur_tier = cur.get("confidence_tier", "")
        base_tier = base.get("confidence_tier", "")
        if cur_tier != base_tier:
            tier_changes.append({
                "primary_xref": c_pxref,
                "standard_name": cur.get("standard_name", ""),
                "old_tier": base_tier,
                "new_tier": cur_tier,
            })
        cur_q = cur.get("overall_quality", "")
        base_q = base.get("overall_quality", "")
        if cur_q != base_q:
            quality_changes.append({
                "primary_xref": c_pxref,
                "standard_name": cur.get("standard_name", ""),
                "old_quality": base_q,
                "new_quality": cur_q,
            })
        cur_type = _concept_biolink_category(cur)
        base_type = _concept_biolink_category(base)
        if cur_type != base_type:
            disease_type_changes.append({
                "primary_xref": c_pxref,
                "standard_name": cur.get("standard_name", ""),
                "old_category": base_type,
                "new_category": cur_type,
            })

    # Compare ncats_id-matched pairs
    for b_pxref, c_pxref in identity_matched_pairs:
        _compare_pair(b_pxref, c_pxref)

    # Compare pxref-matched concepts not already covered by ncats_id matching
    for pxref in shared_pxrefs:
        _compare_pair(pxref, pxref)

    # Edge diffs
    current_edge_keys: set[tuple[str, str, str]] = set()
    for pxref, edges in current.edges_by_pxref.items():
        for e in edges:
            current_edge_keys.add(_xref_edge_key(pxref, e))
    baseline_edge_keys: set[tuple[str, str, str]] = set()
    for pxref, edges in baseline.edges_by_pxref.items():
        for e in edges:
            baseline_edge_keys.add(_xref_edge_key(pxref, e))

    added_edge_keys = current_edge_keys - baseline_edge_keys
    removed_edge_keys = baseline_edge_keys - current_edge_keys
    edges_added = len(added_edge_keys)
    edges_removed = len(removed_edge_keys)
    edges_added_by_namespace = Counter(key[2] or "unknown" for key in added_edge_keys)
    edges_removed_by_namespace = Counter(key[2] or "unknown" for key in removed_edge_keys)

    baseline_type_dist = _concept_type_distribution(baseline)
    current_type_dist = _concept_type_distribution(current)
    baseline_ns_dist = _edge_namespace_distribution(baseline)
    current_ns_dist = _edge_namespace_distribution(current)

    added_by_category = Counter(
        _concept_biolink_category(current.concepts_by_pxref[p])
        for p in added_pxrefs
    )
    removed_by_category = Counter(
        _concept_biolink_category(baseline.concepts_by_pxref[p])
        for p in removed_pxrefs
    )

    # Decision counts
    cur_resolved = sum(
        1 for decs in current.decisions_by_pxref.values()
        for d in decs if d.get("status") == "resolved"
    )
    base_resolved = sum(
        1 for decs in baseline.decisions_by_pxref.values()
        for d in decs if d.get("status") == "resolved"
    )
    decisions_resolved = cur_resolved - base_resolved

    all_added_concepts = [
        {
            "primary_xref": p,
            "standard_name": current.concepts_by_pxref[p].get("standard_name", ""),
            "disease_type": _concept_biolink_category(current.concepts_by_pxref[p]),
        }
        for p in sorted(added_pxrefs)
    ]
    added_concepts = all_added_concepts[:500]

    removed_concepts: list[dict[str, str]] = []
    legacy_source_only_rows: list[dict[str, str]] = []
    other_removed_rows: list[dict[str, str]] = []
    for p in sorted(removed_pxrefs):
        base_concept = baseline.concepts_by_pxref[p]
        name = base_concept.get("standard_name", "")

        # Classify removal reason
        base_edges = baseline.edges_by_pxref.get(p, [])
        base_xref_ids = [e.get("xref_id", "") for e in base_edges if e.get("xref_id", "")]

        # Check if xrefs moved to another concept (merged)
        new_owners: set[str] = set()
        for xid in base_xref_ids:
            new_owners.update(current_xref_owners.get(xid, set()))
        new_owners.discard(p)
        previous_xref_owners = baseline_xref_owners.get(p, set()) - {p}
        source_only_fallback = (
            base_concept.get("harmonized_to_ontology", "").lower() == "false"
            or "fallback" in base_concept.get("primary_xref_grounding_type", "").lower()
        )

        # Check if all baseline xrefs were obsolete
        all_obsolete = (
            len(base_edges) > 0
            and all(
                e.get("xref_is_obsolete", "").lower() in ("true", "1")
                for e in base_edges
                if e.get("xref_id", "")
            )
        )

        # Check if baseline had decisions that led to removal
        base_decisions = baseline.decisions_by_pxref.get(p, [])
        had_omit_decision = any(
            "omit" in (
                d.get("auto_decision", "")
                or d.get("resolution", "")
                or d.get("decision", "")
            ).lower()
            or "drop" in (
                d.get("auto_decision", "")
                or d.get("resolution", "")
                or d.get("decision", "")
            ).lower()
            for d in base_decisions
        )

        bucket = "other_removed"
        if new_owners:
            merged_into = sorted(new_owners)[:3]
            reason = f"Merged into {', '.join(merged_into)}"
            bucket = "retired_concept"
        elif p in current_xref_owners:
            absorbed_into = sorted(current_xref_owners[p])[:3]
            reason = f"Now represented as an xref under {', '.join(absorbed_into)}"
            bucket = "retired_concept"
        elif all_obsolete:
            reason = "Removed because previous xrefs are obsolete"
            bucket = "retired_concept"
        elif had_omit_decision:
            reason = "Removed by QC review decision"
            bucket = "retired_concept"
        elif source_only_fallback and previous_xref_owners:
            reason = "Older single-source export row; not a harmonized concept in the current release"
            bucket = "legacy_source_only"
        elif source_only_fallback:
            reason = "Older single-source export row; not a harmonized concept in the current release"
            bucket = "legacy_source_only"
        elif previous_xref_owners:
            reason = "Previous xref evidence is not retained in the current release"
        elif len(base_xref_ids) == 0:
            reason = "Older row had no exported xref evidence"
        else:
            reason = "Source evidence is not present in the current release"

        item = {
            "primary_xref": p,
            "standard_name": name,
            "disease_type": _concept_biolink_category(base_concept),
            "reason": reason,
        }
        if bucket == "retired_concept":
            removed_concepts.append(item)
        elif bucket == "legacy_source_only":
            legacy_source_only_rows.append(item)
        else:
            other_removed_rows.append(item)

    baseline_version = baseline.manifest.get("pipeline_version", "unknown")
    current_version = current.manifest.get("pipeline_version", "unknown")

    sorted_rekeyed = sorted(rekeyed_concepts, key=lambda x: x["old_primary_xref"])
    sorted_primary_grounding_changes = sorted(
        primary_grounding_changes,
        key=lambda x: x["old_primary_xref"],
    )
    hierarchy_edges_baseline = sum(len(v) for v in baseline.hierarchy_by_child.values())
    hierarchy_edges_current = sum(len(v) for v in current.hierarchy_by_child.values())
    clinical_descendant_edges_baseline = sum(
        len(v) for v in baseline.clinical_descendants_by_pxref.values()
    )
    clinical_descendant_edges_current = sum(
        len(v) for v in current.clinical_descendants_by_pxref.values()
    )
    gene_associations_baseline = sum(len(v) for v in baseline.associations_by_ncats_id.values())
    gene_associations_current = sum(len(v) for v in current.associations_by_ncats_id.values())
    review_records_baseline = sum(len(v) for v in baseline.decisions_by_pxref.values())
    review_records_current = sum(len(v) for v in current.decisions_by_pxref.values())
    data_layer_changes = [
        {
            "layer": "Concepts",
            "old_count": len(baseline.concepts_by_pxref),
            "new_count": len(current.concepts_by_pxref),
        },
        {
            "layer": "Exact xref edges",
            "old_count": len(baseline_edge_keys),
            "new_count": len(current_edge_keys),
        },
        {
            "layer": "Ontology hierarchy edges",
            "old_count": hierarchy_edges_baseline,
            "new_count": hierarchy_edges_current,
        },
        {
            "layer": "Clinical descendant edges",
            "old_count": clinical_descendant_edges_baseline,
            "new_count": clinical_descendant_edges_current,
        },
        {
            "layer": "Disease-gene associations",
            "old_count": gene_associations_baseline,
            "new_count": gene_associations_current,
        },
        {
            "layer": "Xref labels",
            "old_count": len(baseline.labels_by_xref_id),
            "new_count": len(current.labels_by_xref_id),
        },
        {
            "layer": "Review registry records",
            "old_count": review_records_baseline,
            "new_count": review_records_current,
        },
    ]
    for row in data_layer_changes:
        row["delta"] = row["new_count"] - row["old_count"]

    category_delta = _distribution_delta(baseline_type_dist, current_type_dist, "category")
    namespace_delta = _distribution_delta(baseline_ns_dist, current_ns_dist, "namespace")
    source_version_changes = _disease_source_version_changes(baseline, current)

    return {
        "summary": {
            "concepts_added": len(added_pxrefs),
            "concepts_removed": len(removed_concepts),
            "concepts_rekeyed": len(rekeyed_concepts),
            "primary_grounding_changes": len(primary_grounding_changes),
            "raw_primary_rows_added": len(raw_added_pxrefs),
            "raw_primary_rows_absent": len(raw_removed_pxrefs),
            "previous_primary_ids_now_xrefs": len(previous_primary_represented_as_xref),
            "current_primary_rows_explained_by_previous_xrefs": len(
                current_primary_explained_by_previous_xref & raw_added_pxrefs
            ),
            "legacy_source_only_removed": len(legacy_source_only_rows),
            "other_removed_rows": len(other_removed_rows),
            "tier_changes": len(tier_changes),
            "quality_changes": len(quality_changes),
            "disease_type_changes": len(disease_type_changes),
            "edges_added": edges_added,
            "edges_removed": edges_removed,
            "hierarchy_edges_delta": hierarchy_edges_current - hierarchy_edges_baseline,
            "clinical_descendant_edges_added": max(
                0,
                clinical_descendant_edges_current - clinical_descendant_edges_baseline,
            ),
            "gene_associations_added": max(0, gene_associations_current - gene_associations_baseline),
            "decisions_resolved": max(0, decisions_resolved),
            "source_namespaces_added": sum(1 for ns, count in current_ns_dist.items() if count and not baseline_ns_dist.get(ns)),
            "source_namespaces_removed": sum(1 for ns, count in baseline_ns_dist.items() if count and not current_ns_dist.get(ns)),
        },
        "data_layer_changes": data_layer_changes,
        "added_concepts": added_concepts,
        "removed_concepts": removed_concepts[:500],
        "rekeyed_concepts": sorted_rekeyed[:500],
        "primary_grounding_changes": sorted_primary_grounding_changes[:500],
        "legacy_source_only_rows": legacy_source_only_rows[:200],
        "other_removed_rows": other_removed_rows[:200],
        "tier_changes": tier_changes[:500],
        "quality_changes": quality_changes[:500],
        "disease_type_changes": disease_type_changes[:500],
        "truncation": {
            "added_concepts": {"total": len(all_added_concepts), "shown": len(added_concepts)},
            "removed_concepts": {"total": len(removed_concepts), "shown": min(len(removed_concepts), 500)},
            "rekeyed_concepts": {"total": len(rekeyed_concepts), "shown": min(len(rekeyed_concepts), 500)},
            "primary_grounding_changes": {
                "total": len(primary_grounding_changes),
                "shown": min(len(primary_grounding_changes), 500),
            },
            "legacy_source_only_rows": {"total": len(legacy_source_only_rows), "shown": min(len(legacy_source_only_rows), 200)},
            "other_removed_rows": {"total": len(other_removed_rows), "shown": min(len(other_removed_rows), 200)},
            "tier_changes": {"total": len(tier_changes), "shown": min(len(tier_changes), 500)},
            "quality_changes": {"total": len(quality_changes), "shown": min(len(quality_changes), 500)},
            "disease_type_changes": {"total": len(disease_type_changes), "shown": min(len(disease_type_changes), 500)},
        },
        "biolink_category_distribution": {
            "baseline": dict(baseline_type_dist.most_common()),
            "current": dict(current_type_dist.most_common()),
            "delta": category_delta,
            "added": dict(added_by_category.most_common()),
            "removed": dict(removed_by_category.most_common()),
        },
        "source_namespace_distribution": {
            "baseline": dict(baseline_ns_dist.most_common()),
            "current": dict(current_ns_dist.most_common()),
            "delta": namespace_delta,
            "edges_added": dict(edges_added_by_namespace.most_common()),
            "edges_removed": dict(edges_removed_by_namespace.most_common()),
        },
        "source_version_changes": source_version_changes,
        "change_drivers": _disease_change_drivers(
            source_version_changes,
            category_delta,
            namespace_delta,
            data_layer_changes,
        ),
        "change_driver_note": "Source attribution is inferred from source-version changes plus app-graph namespace and layer deltas. NodeNorm is shown as validator/canonicalization context, not as an asserting disease source.",
        "baseline_version": baseline_version,
        "current_version": current_version,
    }


# ---------------------------------------------------------------------------
# Concept Neighborhood Expansion
# ---------------------------------------------------------------------------

def find_concept_neighbors(
    data: DiseaseGraphData,
    pxref: str,
    max_neighbors: int = 10,
) -> list[str]:
    """Find concepts that share xref IDs with the given concept."""
    xref_to_pxrefs = _ensure_xref_index(data)

    # Get all xref_ids for this concept
    my_xrefs = {e.get("xref_id", "") for e in data.edges_by_pxref.get(pxref, []) if e.get("xref_id")}

    # Find other concepts sharing xrefs, ranked by overlap count
    overlap: Counter[str] = Counter()
    for xid in my_xrefs:
        for other in xref_to_pxrefs.get(xid, set()):
            if other != pxref:
                overlap[other] += 1

    return [px for px, _ in overlap.most_common(max_neighbors)]


# ---------------------------------------------------------------------------
# Provenance Chain
# ---------------------------------------------------------------------------

def build_provenance_chain(
    data: DiseaseGraphData,
    pxref: str,
) -> list[dict[str, Any]]:
    """Build ordered provenance steps for a concept."""
    if pxref not in data.concepts_by_pxref:
        return []

    concept = data.concepts_by_pxref[pxref]
    steps: list[dict[str, Any]] = []

    # Step 1: Source assertions (one per xref edge)
    for edge in data.edges_by_pxref.get(pxref, []):
        steps.append({
            "step": "source_assertion",
            "source": _normalize_ns(edge.get("xref_namespace", "")),
            "xref": edge.get("xref_id", ""),
            "match_type": edge.get("match_type", ""),
            "source_asserted": edge.get("source_asserted", ""),
        })

    # Step 2: Confidence scoring (per xref with confidence)
    scored_xrefs: set[str] = set()
    for edge in data.edges_by_pxref.get(pxref, []):
        xid = edge.get("xref_id", "")
        conf = edge.get("equivalence_confidence", "") or edge.get("xref_confidence", "")
        legacy_conf = edge.get("xref_confidence", "")
        if xid and conf and xid not in scored_xrefs:
            scored_xrefs.add(xid)
            steps.append({
                "step": "confidence_scoring",
                "xref": xid,
                "scores": {
                    "recommended_confidence": conf,
                    "equivalence_confidence": edge.get("equivalence_confidence", ""),
                    "relation_confidence": edge.get("relation_confidence", ""),
                    "xref_confidence": legacy_conf,
                    "mapping_confidence": edge.get("mapping_confidence", ""),
                    "label_confidence": edge.get("label_confidence", ""),
                    "source_support": edge.get("source_support", ""),
                    "validation_support": edge.get("validation_support", ""),
                    "corroboration_score": edge.get("corroboration_score", ""),
                    "structural_confidence": edge.get("structural_confidence", ""),
                },
                "evidence_tier": edge.get("evidence_tier", ""),
                "review_tier": edge.get("review_tier", ""),
                "score_rationale": edge.get("score_rationale", ""),
            })

    # Step 3: NodeNorm validation
    nn_status = _nodenorm_status(concept)
    if nn_status:
        steps.append({
            "step": "nodenorm_validation",
            "status": nn_status,
            "canonical": concept.get("nodenorm_canonical_curie", ""),
            "canonical_label": concept.get("nodenorm_canonical_label", ""),
        })

    # Step 4: Conflicts and decisions
    for dec in data.decisions_by_pxref.get(pxref, []):
        steps.append({
            "step": "conflict_detected",
            "type": dec.get("decision_type", ""),
            "xref": dec.get("xref_id", ""),
            "decision_source": dec.get("decision_source", ""),
            "priority": dec.get("priority", ""),
        })
        auto_dec = dec.get("auto_decision", "")
        resolution = dec.get("resolution", "")
        if auto_dec or resolution:
            steps.append({
                "step": "auto_decision" if auto_dec else "manual_resolution",
                "decision": auto_dec or resolution,
                "confidence": dec.get("auto_confidence", ""),
                "rationale": dec.get("auto_rationale", "") or dec.get("resolution_detail", ""),
                "xref": dec.get("xref_id", ""),
            })

    # Step 5: Final resolution status
    steps.append({
        "step": "resolution",
        "status": concept.get("confidence_tier", ""),
        "overall_quality": concept.get("overall_quality", ""),
        "n_sources": concept.get("n_sources", ""),
    })

    return steps


# Names of files that can be served from the app_graph directory.
DOWNLOADABLE_FILES = frozenset({
    "disease_concepts.tsv",
    "disease_xref_edges.tsv",
    "disease_hierarchy_edges.tsv",
    "disease_clinical_descendant_edges.tsv",
    "xref_labels.tsv",
    "review_decisions.tsv",
    "disease_gene_associations.tsv",
    "manifest.json",
})
