from __future__ import annotations

import json
import sqlite3
import threading
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook

try:
    from rdkit import Chem
    from rdkit.Chem import inchi
except ImportError:  # pragma: no cover - depends on QA runtime image
    Chem = None
    inchi = None


QA_BROWSER_DIR = Path(__file__).resolve().parent
MULTIRAMP_WORKBOOK = QA_BROWSER_DIR / "data" / "ramp" / "20260604_final_all_metabolites_ramp_xrefs_multRamp.xlsx"
RAMP_DIAGNOSIS_OPTIONS = [
    {
        "value": "stereo_isomer_can_group",
        "label": "Stereo isomer",
        "decision": "can_group",
        "decision_label": "Can be grouped",
    },
    {
        "value": "positional_isomer_can_group",
        "label": "Positional isomer",
        "decision": "can_group",
        "decision_label": "Can be grouped",
    },
    {
        "value": "salt_included_can_group",
        "label": "Salt included",
        "decision": "can_group",
        "decision_label": "Can be grouped",
    },
    {
        "value": "synonym_names_can_group",
        "label": "Synonym names",
        "decision": "can_group",
        "decision_label": "Can be grouped",
    },
    {
        "value": "distinct_compound_should_not_group",
        "label": "Distinct compound",
        "decision": "should_not_group",
        "decision_label": "Should not be grouped",
    },
    {
        "value": "modified_metabolite_should_not_group",
        "label": "Modified metabolite",
        "decision": "should_not_group",
        "decision_label": "Should not be grouped",
    },
    {
        "value": "other_can_group",
        "label": "Other",
        "decision": "can_group",
        "decision_label": "Can be grouped",
    },
    {
        "value": "other_should_not_group",
        "label": "Other",
        "decision": "should_not_group",
        "decision_label": "Should not be grouped",
    },
]
_DIAGNOSIS_OPTIONS_BY_VALUE = {option["value"]: option for option in RAMP_DIAGNOSIS_OPTIONS}
_diagnosis_file: Path | None = None
_diagnosis_lock = threading.Lock()


def parse_ramp_ids(raw: str) -> list[str]:
    ids = []
    for token in (raw or "").replace("|", " ").replace(",", " ").split():
        normalized = token.strip().upper()
        if normalized.startswith("RAMP_C_") and normalized not in ids:
            ids.append(normalized)
    return ids


def ramp_diagnosis_case_key(ramp_ids: list[str]) -> str:
    return "|".join(sorted(set(ramp_ids)))


def set_ramp_diagnosis_file(path: str | None) -> None:
    global _diagnosis_file
    _diagnosis_file = Path(path) if path else None
    if _diagnosis_file and not _diagnosis_file.exists():
        _diagnosis_file.parent.mkdir(parents=True, exist_ok=True)
        _diagnosis_file.write_text("[]")


def ramp_diagnosis_enabled() -> bool:
    return _diagnosis_file is not None


def _read_ramp_diagnoses() -> list[dict[str, Any]]:
    if not _diagnosis_file or not _diagnosis_file.exists():
        return []
    with _diagnosis_lock:
        try:
            return json.loads(_diagnosis_file.read_text() or "[]")
        except (OSError, json.JSONDecodeError):
            return []


def _write_ramp_diagnoses(entries: list[dict[str, Any]]) -> None:
    if not _diagnosis_file:
        raise HTTPException(status_code=500, detail="RaMP diagnosis storage is not configured.")
    with _diagnosis_lock:
        _diagnosis_file.parent.mkdir(parents=True, exist_ok=True)
        _diagnosis_file.write_text(json.dumps(entries, indent=2, default=str))


def get_ramp_diagnoses(ramp_ids: list[str]) -> list[dict[str, Any]]:
    case_key = ramp_diagnosis_case_key(ramp_ids)
    if not case_key:
        return []
    return [
        entry for entry in _read_ramp_diagnoses()
        if entry.get("case_key") == case_key
    ]


def add_ramp_diagnosis(
    *,
    ramp_ids: list[str],
    diagnosis: str,
    reviewer: str,
    note: str = "",
    workbook_row: int | None = None,
    selected_ramp_ids: list[str] | None = None,
) -> dict[str, Any]:
    if diagnosis not in _DIAGNOSIS_OPTIONS_BY_VALUE:
        raise HTTPException(status_code=400, detail="Unknown diagnosis category.")
    clean_reviewer = reviewer.strip()
    if not clean_reviewer:
        raise HTTPException(status_code=400, detail="Reviewer name is required.")
    case_key = ramp_diagnosis_case_key(ramp_ids)
    if not case_key:
        raise HTTPException(status_code=400, detail="Provide at least one RAMP_C_* ID.")
    normalized_case_ids = sorted(set(ramp_ids))
    normalized_selected_ids = sorted(set(selected_ramp_ids or ramp_ids))
    if not normalized_selected_ids:
        raise HTTPException(status_code=400, detail="Select at least one RaMP ID for this diagnosis.")
    unknown_ids = sorted(set(normalized_selected_ids) - set(normalized_case_ids))
    if unknown_ids:
        raise HTTPException(status_code=400, detail=f"Selected RaMP IDs are not in this case: {', '.join(unknown_ids)}")

    option = _DIAGNOSIS_OPTIONS_BY_VALUE[diagnosis]
    entry = {
        "id": str(uuid.uuid4()),
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "case_key": case_key,
        "ramp_ids": normalized_case_ids,
        "selected_ramp_ids": normalized_selected_ids,
        "workbook_row": workbook_row,
        "reviewer": clean_reviewer,
        "diagnosis": diagnosis,
        "diagnosis_label": option["label"],
        "decision": option["decision"],
        "decision_label": option["decision_label"],
        "note": note.strip(),
    }
    entries = _read_ramp_diagnoses()
    entries.append(entry)
    _write_ramp_diagnoses(entries)
    return entry


def delete_ramp_diagnosis(entry_id: str) -> bool:
    if not entry_id:
        return False
    entries = _read_ramp_diagnoses()
    kept_entries = [entry for entry in entries if entry.get("id") != entry_id]
    if len(kept_entries) == len(entries):
        return False
    _write_ramp_diagnoses(kept_entries)
    return True


def summarize_ramp_diagnoses(entries: list[dict[str, Any]]) -> dict[str, Any]:
    can_group = [entry for entry in entries if entry.get("decision") == "can_group"]
    should_not_group = [entry for entry in entries if entry.get("decision") == "should_not_group"]
    reviewers = sorted({entry.get("reviewer") for entry in entries if entry.get("reviewer")})
    return {
        "total": len(entries),
        "can_group_count": len(can_group),
        "should_not_group_count": len(should_not_group),
        "reviewers": reviewers,
        "entries": entries,
    }


def attach_ramp_diagnosis_summaries(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    entries_by_case: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in _read_ramp_diagnoses():
        case_key = entry.get("case_key")
        if case_key:
            entries_by_case[case_key].append(entry)

    for row in rows:
        case_key = ramp_diagnosis_case_key(parse_ramp_ids(row.get("finalRampIds") or ""))
        row["curationSummary"] = summarize_ramp_diagnoses(entries_by_case.get(case_key, []))
    return rows


def parse_hit_count(value: Any) -> int:
    if value is None:
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return 0


def fetch_analytes(db_path: Path, ramp_ids: list[str]) -> list[dict[str, Any]]:
    placeholders = ",".join("?" for _ in ramp_ids)
    query = f"""
        SELECT rampId, type, common_name
        FROM analyte
        WHERE rampId IN ({placeholders})
        ORDER BY rampId
    """
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        return [dict(row) for row in con.execute(query, ramp_ids)]


def fetch_source_rows(db_path: Path, ramp_ids: list[str]) -> list[dict[str, Any]]:
    placeholders = ",".join("?" for _ in ramp_ids)
    query = f"""
        SELECT
            s.rampId,
            a.common_name AS rampName,
            s.sourceId,
            s.IDtype,
            s.dataSource,
            s.commonName AS sourceName,
            s.priorityHMDBStatus,
            s.pathwayCount
        FROM source s
        JOIN analyte a ON a.rampId = s.rampId
        WHERE s.rampId IN ({placeholders})
        ORDER BY s.rampId, s.dataSource, s.IDtype, s.sourceId
    """
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        return [dict(row) for row in con.execute(query, ramp_ids)]


def fetch_chem_props(db_path: Path, ramp_ids: list[str]) -> list[dict[str, Any]]:
    placeholders = ",".join("?" for _ in ramp_ids)
    query = f"""
        SELECT
            ramp_id,
            chem_data_source,
            chem_source_id,
            mw,
            monoisotop_mass,
            common_name,
            mol_formula,
            iso_smiles,
            inchi_key
        FROM chem_props
        WHERE ramp_id IN ({placeholders})
        ORDER BY ramp_id, chem_data_source, chem_source_id
    """
    with sqlite3.connect(db_path) as con:
        con.row_factory = sqlite3.Row
        return [dict(row) for row in con.execute(query, ramp_ids)]


def source_node_id(source_id: str) -> str:
    return f"source::{source_id}"


def inchi_key_node_id(inchi_key: str) -> str:
    return f"inchikey::{inchi_key}"


def stereo_free_node_id(inchi_key: str) -> str:
    return f"stereo-free::{inchi_key}"


def stereo_free_inchi_key(smiles: str | None) -> tuple[str, str, str]:
    if not smiles:
        return "", "", ""
    if Chem is None or inchi is None:
        return "", "", "RDKit is not installed"
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return "", "", "RDKit could not parse SMILES"
    Chem.RemoveStereochemistry(mol)
    no_stereo_smiles = Chem.MolToSmiles(mol, isomericSmiles=False)
    return inchi.MolToInchiKey(mol), no_stereo_smiles, ""


def format_number(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.6g}"
    except (TypeError, ValueError):
        return str(value)


def join_values(values: set[str]) -> str:
    return ", ".join(sorted(value for value in values if value))


def finalize_sets(data: dict[str, Any]) -> None:
    for key, value in list(data.items()):
        if isinstance(value, set):
            data[key] = join_values(value)


def build_ramp_graph_payload(db_path: Path, ramp_ids: list[str], show_individual_ids: bool = False) -> dict[str, Any]:
    if not db_path.exists():
        raise HTTPException(status_code=500, detail=f"RaMP SQLite database not found: {db_path}")
    if not ramp_ids:
        raise HTTPException(status_code=400, detail="Provide at least one RAMP_C_* ID.")

    analytes = fetch_analytes(db_path, ramp_ids)
    found = {row["rampId"] for row in analytes}
    missing = [ramp_id for ramp_id in ramp_ids if ramp_id not in found]
    source_rows = fetch_source_rows(db_path, ramp_ids)
    chem_rows = fetch_chem_props(db_path, ramp_ids)
    chem_by_ramp_source: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in chem_rows:
        chem_by_ramp_source[(row["ramp_id"], row["chem_source_id"])].append(row)

    ramp_nodes: dict[str, dict[str, Any]] = {}
    source_nodes: dict[str, dict[str, Any]] = {}
    inchi_key_nodes: dict[str, dict[str, Any]] = {}
    stereo_free_nodes: dict[str, dict[str, Any]] = {}
    edges: dict[str, dict[str, Any]] = {}
    inchi_edges: dict[str, dict[str, Any]] = {}
    stereo_free_edges: dict[str, dict[str, Any]] = {}
    stereo_free_errors: set[str] = set()

    for row in source_rows:
        ramp_nodes.setdefault(row["rampId"], {
            "data": {
                "id": row["rampId"],
                "label": row["rampId"],
                "name": row["rampName"],
                "kind": "ramp",
                "sourceIds": set(),
                "dataSources": set(),
                "sourceNames": set(),
                "molecularWeights": set(),
                "monoisotopicMasses": set(),
                "formulas": set(),
                "smiles": set(),
                "inchiKeys": set(),
            },
            "classes": "ramp",
        })
        source_id = row["sourceId"]
        ramp_nodes[row["rampId"]]["data"]["sourceIds"].add(source_id)
        ramp_nodes[row["rampId"]]["data"]["dataSources"].add(row["dataSource"] or "")
        ramp_nodes[row["rampId"]]["data"]["sourceNames"].add(row["sourceName"] or "")

        if show_individual_ids:
            source_nodes.setdefault(source_id, {
                "data": {
                    "id": source_node_id(source_id),
                    "label": source_id,
                    "sourceId": source_id,
                    "idType": row["IDtype"],
                    "kind": "source",
                    "names": set(),
                    "dataSources": set(),
                    "molecularWeights": set(),
                    "monoisotopicMasses": set(),
                    "formulas": set(),
                    "smiles": set(),
                    "inchiKeys": set(),
                },
                "classes": f"source {str(row['IDtype']).lower()}",
            })
            source_nodes[source_id]["data"]["names"].add(row["sourceName"] or "")
            source_nodes[source_id]["data"]["dataSources"].add(row["dataSource"] or "")

            edge_id = f"edge::{row['rampId']}::{source_id}"
            edges.setdefault(edge_id, {
                "data": {
                    "id": edge_id,
                    "source": row["rampId"],
                    "target": source_node_id(source_id),
                    "sourceId": source_id,
                    "label": set(),
                },
            })
            edges[edge_id]["data"]["label"].add(row["dataSource"] or "")

        for chem_row in chem_by_ramp_source.get((row["rampId"], source_id), []):
            chem_target_data = [ramp_nodes[row["rampId"]]["data"]]
            if show_individual_ids:
                chem_target_data.append(source_nodes[source_id]["data"])
            for data in chem_target_data:
                data["molecularWeights"].add(format_number(chem_row["mw"]))
                data["monoisotopicMasses"].add(format_number(chem_row["monoisotop_mass"]))
                data["formulas"].add(chem_row["mol_formula"] or "")
                data["smiles"].add(chem_row["iso_smiles"] or "")
                data["inchiKeys"].add(chem_row["inchi_key"] or "")
            inchi_key = chem_row["inchi_key"] or ""
            if not inchi_key:
                continue
            inchi_key_nodes.setdefault(inchi_key, {
                "data": {
                    "id": inchi_key_node_id(inchi_key),
                    "label": inchi_key,
                    "inchiKey": inchi_key,
                    "kind": "inchiKey",
                    "names": set(),
                    "formulas": set(),
                    "molecularWeights": set(),
                },
                "classes": "inchi-key",
            })
            inchi_key_nodes[inchi_key]["data"]["names"].add(chem_row["common_name"] or "")
            inchi_key_nodes[inchi_key]["data"]["formulas"].add(chem_row["mol_formula"] or "")
            inchi_key_nodes[inchi_key]["data"]["molecularWeights"].add(format_number(chem_row["mw"]))
            inchi_source_id = source_node_id(source_id) if show_individual_ids else row["rampId"]
            inchi_edge_id = f"inchi::{inchi_source_id}::{inchi_key}"
            inchi_edges.setdefault(inchi_edge_id, {
                "data": {
                    "id": inchi_edge_id,
                    "source": inchi_source_id,
                    "target": inchi_key_node_id(inchi_key),
                    "sourceId": source_id,
                    "label": set(),
                },
                "classes": "inchi-key-edge",
            })
            inchi_edges[inchi_edge_id]["data"]["label"].add(row["dataSource"] or "InChIKey")
            stereo_free_key, no_stereo_smiles, stereo_error = stereo_free_inchi_key(chem_row["iso_smiles"])
            if stereo_error:
                stereo_free_errors.add(stereo_error)
            if not stereo_free_key:
                continue
            stereo_free_nodes.setdefault(stereo_free_key, {
                "data": {
                    "id": stereo_free_node_id(stereo_free_key),
                    "label": stereo_free_key,
                    "stereoFreeInchiKey": stereo_free_key,
                    "kind": "stereoFreeKey",
                    "names": set(),
                    "formulas": set(),
                    "smiles": set(),
                    "noStereoSmiles": set(),
                    "molecularWeights": set(),
                    "monoisotopicMasses": set(),
                    "inchiKeys": set(),
                },
                "classes": "stereo-free-key",
            })
            stereo_free_nodes[stereo_free_key]["data"]["names"].add(chem_row["common_name"] or "")
            stereo_free_nodes[stereo_free_key]["data"]["formulas"].add(chem_row["mol_formula"] or "")
            stereo_free_nodes[stereo_free_key]["data"]["smiles"].add(chem_row["iso_smiles"] or "")
            stereo_free_nodes[stereo_free_key]["data"]["noStereoSmiles"].add(no_stereo_smiles)
            stereo_free_nodes[stereo_free_key]["data"]["molecularWeights"].add(format_number(chem_row["mw"]))
            stereo_free_nodes[stereo_free_key]["data"]["monoisotopicMasses"].add(format_number(chem_row["monoisotop_mass"]))
            stereo_free_nodes[stereo_free_key]["data"]["inchiKeys"].add(inchi_key)
            stereo_edge_id = f"stereo-free::{inchi_key}::{stereo_free_key}"
            stereo_free_edges.setdefault(stereo_edge_id, {
                "data": {
                    "id": stereo_edge_id,
                    "source": inchi_key_node_id(inchi_key),
                    "target": stereo_free_node_id(stereo_free_key),
                    "inchiKey": inchi_key,
                    "stereoFreeInchiKey": stereo_free_key,
                    "noStereoSmiles": no_stereo_smiles,
                    "label": "no-stereo",
                },
                "classes": "stereo-free-edge",
            })

    elements = [
        *ramp_nodes.values(),
        *source_nodes.values(),
        *inchi_key_nodes.values(),
        *stereo_free_nodes.values(),
        *edges.values(),
        *inchi_edges.values(),
        *stereo_free_edges.values(),
    ]
    for element in elements:
        finalize_sets(element.get("data") or {})

    return {
        "queryRampIds": ramp_ids,
        "missingRampIds": missing,
        "stats": {
            "rampCount": len(ramp_nodes),
            "sourceCount": len(source_nodes),
            "inchiKeyCount": len(inchi_key_nodes),
            "stereoFreeKeyCount": len(stereo_free_nodes),
            "rampSourceEdgeCount": len(edges),
            "inchiKeyEdgeCount": len(inchi_edges),
            "stereoFreeEdgeCount": len(stereo_free_edges),
        },
        "warnings": sorted(stereo_free_errors),
        "analytes": analytes,
        "sourceRows": source_rows,
        "chemRows": chem_rows,
        "elements": elements,
    }


def load_multiramp_rows(limit: int = 500) -> list[dict[str, Any]]:
    if not MULTIRAMP_WORKBOOK.exists():
        return []

    workbook = load_workbook(MULTIRAMP_WORKBOOK, read_only=True, data_only=True)
    worksheet = workbook["concatenated_xrefs"]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(rows)]
    indexes = {name: idx for idx, name in enumerate(header)}
    required = ["Standardized name", "final_RaMP_ids", "final_RaMP_hit_count"]
    if any(name not in indexes for name in required):
        workbook.close()
        return []

    output: list[dict[str, Any]] = []
    for workbook_row_index, row in enumerate(rows, start=2):
        hit_count = parse_hit_count(row[indexes["final_RaMP_hit_count"]])
        if hit_count <= 1:
            continue
        ramp_ids = parse_ramp_ids(str(row[indexes["final_RaMP_ids"]] or ""))
        if len(ramp_ids) <= 1:
            continue
        item = {
            "workbookRow": workbook_row_index,
            "standardizedName": row[indexes["Standardized name"]] or "",
            "chemicalName": row[indexes.get("CHEMICAL_NAME", -1)] if "CHEMICAL_NAME" in indexes else "",
            "chemId": row[indexes.get("CHEM_ID", -1)] if "CHEM_ID" in indexes else "",
            "metabolonId": row[indexes.get("METABOLON_ID", -1)] if "METABOLON_ID" in indexes else "",
            "finalRampIds": " | ".join(ramp_ids),
            "finalRampHitCount": hit_count,
            "href": "/ramp-id-qa?ids=" + "%20%7C%20".join(ramp_ids),
        }
        output.append(item)
        if len(output) >= limit:
            break
    workbook.close()
    return output


def load_multiramp_export_rows() -> tuple[list[str], list[dict[str, Any]]]:
    if not MULTIRAMP_WORKBOOK.exists():
        return [], []

    workbook = load_workbook(MULTIRAMP_WORKBOOK, read_only=True, data_only=True)
    worksheet = workbook["concatenated_xrefs"]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(rows)]
    indexes = {name: idx for idx, name in enumerate(header)}
    required = ["final_RaMP_ids", "final_RaMP_hit_count"]
    if any(name not in indexes for name in required):
        workbook.close()
        return header, []

    entries_by_case: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in _read_ramp_diagnoses():
        case_key = entry.get("case_key")
        if case_key:
            entries_by_case[case_key].append(entry)

    export_rows: list[dict[str, Any]] = []
    for workbook_row_index, row in enumerate(rows, start=2):
        hit_count = parse_hit_count(row[indexes["final_RaMP_hit_count"]])
        if hit_count <= 1:
            continue
        ramp_ids = parse_ramp_ids(str(row[indexes["final_RaMP_ids"]] or ""))
        if len(ramp_ids) <= 1:
            continue
        workbook_values = {
            name: row[index] if index < len(row) and row[index] is not None else ""
            for name, index in indexes.items()
        }
        base = {
            "workbook_row": workbook_row_index,
            **workbook_values,
        }
        entries = entries_by_case.get(ramp_diagnosis_case_key(ramp_ids), [])
        if not entries:
            export_rows.append(base)
            continue
        for entry in entries:
            export_rows.append({
                **base,
                "curation_id": entry.get("id", ""),
                "curation_timestamp": entry.get("timestamp", ""),
                "curation_reviewer": entry.get("reviewer", ""),
                "curation_decision": entry.get("decision", ""),
                "curation_decision_label": entry.get("decision_label", ""),
                "curation_diagnosis": entry.get("diagnosis", ""),
                "curation_diagnosis_label": entry.get("diagnosis_label", ""),
                "curation_selected_ramp_ids": " | ".join(entry.get("selected_ramp_ids") or entry.get("ramp_ids") or []),
                "curation_note": entry.get("note", ""),
            })

    workbook.close()
    return header, export_rows


def build_multiramp_navigation(raw_ids: str, limit: int = 5000) -> dict[str, Any] | None:
    current_ids = parse_ramp_ids(raw_ids)
    if not current_ids:
        return None
    current_set = set(current_ids)
    rows = load_multiramp_rows(limit=limit)
    current_index = None
    for index, row in enumerate(rows):
        if set(parse_ramp_ids(row["finalRampIds"])) == current_set:
            current_index = index
            break
    if current_index is None:
        return None
    current = rows[current_index]
    return {
        "current": current,
        "previous": rows[current_index - 1] if current_index > 0 else None,
        "next": rows[current_index + 1] if current_index + 1 < len(rows) else None,
        "index": current_index + 1,
        "total": len(rows),
    }
